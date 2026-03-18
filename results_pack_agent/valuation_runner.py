# results_pack_agent/valuation_runner.py
# Invoke the Wally value-chart builder for the results pack ticker.
# This wraps wally/value_chart_builder.py cleanly without touching Bob logic.

from __future__ import annotations

from pathlib import Path
from typing import Optional

from .utils import log

# Optional: import the wally value chart builder.
# This may not be available in all environments (e.g. minimal CI installs).
try:
    from wally.value_chart_builder import build_value_chart as _build_value_chart  # type: ignore[import]
    _WALLY_AVAILABLE = True
except ImportError:
    _build_value_chart = None  # type: ignore[assignment]
    _WALLY_AVAILABLE = False


def build_valuation(
    ticker: str,
    output_folder: Path,
    file_prefix: str,
    dry_run: bool = False,
) -> Optional[str]:
    """Build the Wally value-chart workbook for *ticker*.

    The workbook is saved to *output_folder* with name
    ``{file_prefix}-Value-Chart.xlsx``.

    Returns the local path to the created workbook, or ``None`` on failure.
    If *dry_run* is True, logs the action and returns the would-be path.
    """
    out_path = output_folder / f"{file_prefix}-Value-Chart.xlsx"

    if dry_run:
        log(f"[valuation_runner] [DRY-RUN] Would build value chart → {out_path.name}")
        return str(out_path)

    if not _WALLY_AVAILABLE or _build_value_chart is None:
        log("[valuation_runner] wally.value_chart_builder not available.")
        return _build_empty_shell(out_path, ticker)

    try:
        log(f"[valuation_runner] Building value chart for {ticker} …")
        result_path = _build_value_chart(
            ticker_or_config_path=ticker,
            output_path=str(out_path),
        )
        log(f"[valuation_runner] Value chart saved → {result_path}")
        return result_path
    except Exception as exc:
        log(f"[valuation_runner] Value chart build failed for {ticker}: {exc}")
        # Still create an empty workbook shell so the run folder is complete
        return _build_empty_shell(out_path, ticker)


def _build_empty_shell(out_path: Path, ticker: str) -> Optional[str]:
    """Create a minimal placeholder workbook when the full build fails."""
    try:
        from openpyxl import Workbook  # type: ignore[import]

        wb = Workbook()
        ws = wb.active
        ws.title = "DataRequired"
        ws["A1"] = f"Value chart for {ticker} could not be built automatically."
        ws["A2"] = "Please populate manually or re-run with a valid valuations YAML config."
        wb.save(str(out_path))
        log(f"[valuation_runner] Empty shell workbook saved → {out_path.name}")
        return str(out_path)
    except Exception as exc:
        log(f"[valuation_runner] Could not create shell workbook: {exc}")
        return None
