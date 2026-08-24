#!/usr/bin/env python3
"""Turn the private IRR workbook into a public, money-free decision file.

The workbook holds every buy, its cost, and what it is worth today. This
repository is public, so the workbook itself must never be committed — it is
in .gitignore for that reason.

What Theo actually needs from it survives the money being removed. IRR and
multiple are both scale-invariant: multiply every cash flow in a decision by
any positive constant and neither number moves. So each decision's flows are
normalised so the buy is exactly -1.0, and the IRR solved from them is
identical to the IRR solved from the dollars.

What is published: ticker, which buy it was, the date, the price paid, the
status, each decision's share of total capital committed, and the normalised
flows. What is not: cost, share counts, dividends received, sale proceeds and
current value, in dollars. Price times shares would give the cost back, so
share counts go too.

Run it whenever a new workbook arrives:

    python scripts/export_ledger.py [path/to/workbook.xlsx]
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

SOURCE = Path("data/JM_Decision_Level_IRR.xlsx")
TARGET = Path("data/decisions.json")

DECISION_SHEET = "Decision IRR"
CASHFLOW_SHEET = "Cash Flows"
_DECISION_ID_RE = re.compile(r"^D\d+$")

# Anything below this is rounding noise in a normalised flow, and keeping four
# decimals stops the file from becoming a wall of float dust.
PLACES = 6


def _ticker(symbol: object) -> str:
    text = str(symbol or "").strip().upper()
    if ":" in text:
        text = text.split(":", 1)[1]
    return text.rstrip(".").strip()


def main(argv: list[str]) -> int:
    import openpyxl

    source = Path(argv[1]) if len(argv) > 1 else SOURCE
    if not source.is_file():
        print(f"No workbook at {source}", file=sys.stderr)
        return 1

    book = openpyxl.load_workbook(source, data_only=True, read_only=True)

    flows: dict[str, list[tuple[str, float]]] = {}
    valued_at = ""
    for row in book[CASHFLOW_SHEET].iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not _DECISION_ID_RE.match(str(row[0]).strip()):
            continue
        when, kind, amount = row[2], str(row[3] or "").strip().upper(), row[4]
        if when is None or amount is None:
            continue
        stamp = when.date().isoformat() if hasattr(when, "date") else str(when)[:10]
        flows.setdefault(str(row[0]).strip(), []).append((stamp, float(amount)))
        if kind == "TERMINAL":
            valued_at = max(valued_at, stamp)

    rows = [
        r
        for r in book[DECISION_SHEET].iter_rows(min_row=2, values_only=True)
        if r and r[0] and _DECISION_ID_RE.match(str(r[0]).strip())
    ]
    total_cost = sum(float(r[5] or 0) for r in rows) or 1.0

    counts: dict[str, int] = {}
    decisions = []
    for row in sorted(rows, key=lambda r: (_ticker(r[1]), r[2] or 0)):
        ticker = _ticker(row[1])
        cost = float(row[5] or 0)
        if cost <= 0:
            continue
        counts[ticker] = counts.get(ticker, 0) + 1
        when = row[2]
        decisions.append(
            {
                "ticker": ticker,
                "index": counts[ticker],
                "date": when.date().isoformat() if hasattr(when, "date") else str(when)[:10],
                "price": round(float(row[4] or 0), 4),
                "status": str(row[13] or "OPEN").strip().upper(),
                # Demerger scrip: it arrived, it was not bought. Normalising
                # against a near-zero cost base produces a four-figure
                # "multiple" that means nothing, so it is flagged here and the
                # loader gives it no IRR — the same rule the workbook parser
                # applies to the dollar figures.
                "scrip": cost < 100.0 and float(row[3] or 0) > 100.0,
                # Share of total capital ever committed. A ratio, not an
                # amount — and the number the sizing analysis actually needs.
                "weight": round(cost / total_cost, PLACES),
                "flows": [
                    [stamp, round(amount / cost, PLACES)]
                    for stamp, amount in sorted(flows.get(str(row[0]).strip(), []))
                ],
            }
        )

    payload = {
        "as_at": valued_at,
        "note": (
            "Money-free export of the decision-level IRR workbook. Each "
            "decision's cash flows are normalised so the buy is -1.0, which "
            "leaves IRR and multiple unchanged and leaves no dollar amount "
            "behind. Regenerate with scripts/export_ledger.py."
        ),
        "decisions": decisions,
    }
    TARGET.write_text(json.dumps(payload, indent=1) + "\n")
    print(f"Wrote {TARGET} — {len(decisions)} decisions, valued at {valued_at}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
