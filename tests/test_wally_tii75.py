# tests/test_wally_tii75.py
#
# Tests for the TII75 canonical watchlist and Wally's screening logic.

from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# Ensure repo root is on path so wally package is importable.
_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from wally.watchlist_loader import load_watchlist
from wally.screening import screen_snapshot
from wally.data_fetch import PriceSnapshot, ValuationSnapshot

_TII75_PATH = _REPO_ROOT / "watchlists" / "tii75_watchlist.yaml"

_REQUIRED_TICKERS = {"POOL", "FICO", "CPRT", "2914.T"}


# ---------------------------------------------------------------------------
# TII75 watchlist loading tests
# ---------------------------------------------------------------------------

class TestTII75WatchlistLoading:
    def test_loads_exactly_30_tickers(self):
        wl = load_watchlist(_TII75_PATH, validate_tii75=True)
        assert len(wl.tickers) == 30, (
            f"Expected 30 tickers, got {len(wl.tickers)}: {wl.tickers}"
        )

    def test_pool_is_present(self):
        wl = load_watchlist(_TII75_PATH)
        assert "POOL" in wl.tickers

    def test_fico_is_present(self):
        wl = load_watchlist(_TII75_PATH)
        assert "FICO" in wl.tickers

    def test_cprt_is_present(self):
        wl = load_watchlist(_TII75_PATH)
        assert "CPRT" in wl.tickers

    def test_2914t_is_present(self):
        wl = load_watchlist(_TII75_PATH)
        assert "2914.T" in wl.tickers

    def test_all_required_canonical_members_present(self):
        wl = load_watchlist(_TII75_PATH)
        ticker_set = set(wl.tickers)
        missing = _REQUIRED_TICKERS - ticker_set
        assert not missing, f"Missing required TII75 tickers: {missing}"

    def test_watchlist_name_is_set(self):
        wl = load_watchlist(_TII75_PATH)
        assert wl.name, "Watchlist name should not be empty"

    def test_validate_tii75_raises_on_bad_list(self, tmp_path):
        """validate_tii75=True should raise when the list is incomplete."""
        bad_yaml = tmp_path / "bad_tii75.yaml"
        bad_yaml.write_text("name: TII75 Watchlist\ntickers:\n  - POOL\n  - FICO\n")
        with pytest.raises(ValueError, match="canonical validation"):
            load_watchlist(bad_yaml, validate_tii75=True)


# ---------------------------------------------------------------------------
# Screening unit tests
# ---------------------------------------------------------------------------

class TestScreenSnapshot:
    def _make_snap(self, ticker, current_price, low_52w, high_52w=None):
        return PriceSnapshot(
            ticker=ticker,
            company_name=ticker,
            current_price=current_price,
            low_52w=low_52w,
            high_52w=high_52w if high_52w is not None else current_price * 1.5,
        )

    def test_pool_near_52w_low_is_flagged(self):
        """POOL at 208.78 with 52w low of 203.80 → ~2.44% above low → flagged."""
        snap = self._make_snap("POOL", current_price=208.78, low_52w=203.80, high_52w=353.00)
        result = screen_snapshot(snap, threshold_pct=5.0)

        expected_dist = ((208.78 - 203.80) / 203.80) * 100
        assert abs(result.distance_to_low_pct - expected_dist) < 0.01, (
            f"Expected distance_to_low_pct ≈ {expected_dist:.4f}, got {result.distance_to_low_pct:.4f}"
        )
        assert result.flagged is True, "POOL should be flagged at 2.44% above 52w low"

    def test_msft_far_from_52w_low_not_flagged(self):
        """MSFT at 399.95 with 52w low of 344.79 → >5% above low → not flagged."""
        snap = self._make_snap("MSFT", current_price=399.95, low_52w=344.79)
        result = screen_snapshot(snap, threshold_pct=5.0)

        expected_dist = ((399.95 - 344.79) / 344.79) * 100
        assert abs(result.distance_to_low_pct - expected_dist) < 0.01
        assert result.flagged is False, "MSFT should NOT be flagged when >5% above 52w low"

    def test_exactly_at_threshold_is_flagged(self):
        """Ticker exactly at threshold boundary should be flagged (<=)."""
        snap = self._make_snap("TEST", current_price=105.0, low_52w=100.0)
        result = screen_snapshot(snap, threshold_pct=5.0)
        assert result.flagged is True

    def test_just_above_threshold_not_flagged(self):
        """Ticker just above threshold should NOT be flagged."""
        snap = self._make_snap("TEST", current_price=105.01, low_52w=100.0)
        result = screen_snapshot(snap, threshold_pct=5.0)
        assert result.flagged is False

    def test_invalid_52w_low_returns_unflagged(self):
        """Zero or negative 52w low → unflagged with error."""
        snap = self._make_snap("TEST", current_price=100.0, low_52w=0.0)
        result = screen_snapshot(snap, threshold_pct=5.0)
        assert result.flagged is False
        assert result.error is not None


# ---------------------------------------------------------------------------
# Claude analyst unit tests
# ---------------------------------------------------------------------------

class TestClaudeAnalyst:
    """Unit tests for wally.claude_analyst.analyse_opportunity."""

    def test_returns_none_when_no_api_key(self):
        """analyse_opportunity returns None gracefully when ANTHROPIC_API_KEY is absent."""
        from wally.claude_analyst import analyse_opportunity

        with patch.dict("os.environ", {}, clear=True):
            result = analyse_opportunity(
                ticker="CPRT",
                company_name="Copart Inc.",
                summary={"current_price": 50.0, "low_52w": 49.0, "high_52w": 70.0, "distance_to_low_pct": 2.04},
                reasons=["Trading 2.0% above 52-week low of 49.00"],
            )
        assert result is None

    def test_returns_dict_with_expected_keys_on_success(self):
        """analyse_opportunity returns dict with all five section keys when API call succeeds."""
        from wally.claude_analyst import analyse_opportunity

        fake_block = MagicMock()
        fake_block.type = "text"
        fake_block.text = (
            "VERDICT: Looks interesting.\n\n"
            "BULL CASE: Strong moat.\n\n"
            "BEAR CASE: Market risk.\n\n"
            "WHAT MUST BE TRUE: Earnings stable.\n\n"
            "RECOMMENDATION: Monitor for now."
        )
        fake_response = MagicMock()
        fake_response.content = [fake_block]

        fake_client = MagicMock()
        fake_client.messages.create.return_value = fake_response

        fake_anthropic = MagicMock()
        fake_anthropic.Anthropic.return_value = fake_client

        with patch.dict("os.environ", {"ANTHROPIC_API_KEY": "test-key"}):
            with patch.dict("sys.modules", {"anthropic": fake_anthropic}):
                result = analyse_opportunity(
                    ticker="CPRT",
                    company_name="Copart Inc.",
                    summary={"current_price": 50.0, "low_52w": 49.0},
                    reasons=["Near 52-week low"],
                )

        assert result is not None
        assert set(result.keys()) == {"verdict", "bull_case", "bear_case", "what_must_be_true", "recommendation"}
        assert "interesting" in result["verdict"].lower()
        assert result["recommendation"] == "Monitor for now."


# ---------------------------------------------------------------------------
# Valuation workbook builder unit tests
# ---------------------------------------------------------------------------

class TestValuationWorkbook:
    """Unit tests for wally.valuation_workbook.build_valuation_workbook."""

    def test_creates_xlsx_without_claude_analysis(self, tmp_path):
        """build_valuation_workbook creates a valid .xlsx when claude_analysis is None."""
        from wally.valuation_workbook import build_valuation_workbook

        out = tmp_path / "cprt_value_chart.xlsx"
        summary = {
            "company_name": "Copart Inc.",
            "ticker": "CPRT",
            "current_price": 50.0,
            "low_52w": 49.0,
            "high_52w": 70.0,
            "distance_to_low_pct": 2.04,
            "trailing_pe": 25.0,
            "forward_pe": 22.0,
            "ev_ebitda": 18.0,
            "price_to_sales": 4.5,
            "fcf_yield": 0.04,
            "dividend_yield": None,
        }
        build_valuation_workbook(out, summary=summary, history_rows=[], decision_rows=[], claude_analysis=None)

        assert out.exists()
        import openpyxl
        wb = openpyxl.load_workbook(out)
        assert "Summary Dashboard" in wb.sheetnames
        assert "Decision Framework" in wb.sheetnames
        assert "Claude AI Analysis" not in wb.sheetnames

    def test_creates_claude_sheet_when_analysis_provided(self, tmp_path):
        """build_valuation_workbook includes 'Claude AI Analysis' sheet when analysis is given."""
        from wally.valuation_workbook import build_valuation_workbook

        out = tmp_path / "orly_value_chart.xlsx"
        claude_analysis = {
            "verdict": "Looks interesting.",
            "bull_case": "Strong moat.",
            "bear_case": "Market risk.",
            "what_must_be_true": "Earnings stable.",
            "recommendation": "Monitor for now.",
        }
        build_valuation_workbook(
            out,
            summary={"ticker": "ORLY", "company_name": "O'Reilly Automotive"},
            history_rows=[],
            decision_rows=[],
            claude_analysis=claude_analysis,
        )

        assert out.exists()
        import openpyxl
        wb = openpyxl.load_workbook(out)
        assert "Claude AI Analysis" in wb.sheetnames

    def test_all_expected_sheets_present(self, tmp_path):
        """All six standard sheets are always present in the output workbook."""
        from wally.valuation_workbook import build_valuation_workbook

        out = tmp_path / "adbe_value_chart.xlsx"
        build_valuation_workbook(
            out,
            summary={"ticker": "ADBE", "company_name": "Adobe Inc."},
            history_rows=[{"period": "current", "trailing_pe": 30.0}],
            decision_rows=[{"issue": "Near 52-week low"}],
            claude_analysis=None,
        )

        import openpyxl
        wb = openpyxl.load_workbook(out)
        for expected in [
            "Summary Dashboard",
            "Historical Valuation",
            "Valuation Comparison",
            "Implied Expectations",
            "Critical Review Notes",
            "Decision Framework",
        ]:
            assert expected in wb.sheetnames, f"Missing sheet: {expected}"


# ---------------------------------------------------------------------------
# Integration: fallback workbook path registers range PNG as inline image
# ---------------------------------------------------------------------------

class TestFallbackInlineImage:
    """Verify that _process_watchlist registers the range-chart PNG as an
    inline email image when the fallback valuation workbook is used.

    This ensures TII75 companies (which have no valuations YAML config) show
    a visual chart in the email body, matching the experience for Aussie Tech
    companies that DO have a config (whose value-chart PNG is registered inline).
    """

    def test_range_png_registered_as_inline_image_for_fallback_ticker(self, tmp_path):
        """When a ticker has no valuations config, the fallback workbook is built
        AND the 52-week range chart PNG is added to inline_images so the email
        body shows a chart rather than plain text."""
        import os
        from pathlib import Path
        from unittest.mock import MagicMock, patch

        from wally.data_fetch import PriceSnapshot, ValuationSnapshot

        # ── Mock price and valuation data ───────────────────────────────────
        fake_snap = PriceSnapshot(
            ticker="CPRT",
            company_name="Copart Inc.",
            current_price=33.88,
            low_52w=33.88,
            high_52w=63.84,
        )
        fake_val_snap = ValuationSnapshot(
            trailing_pe=25.0,
            forward_pe=22.0,
            ev_to_ebitda=18.0,
            price_to_sales=4.5,
            fcf_yield=0.04,
            dividend_yield=None,
        )

        # Create a real (but tiny) PNG so path.read_bytes() in send_email succeeds
        range_png_path = tmp_path / "cprt_range.png"
        range_png_path.write_bytes(b"\x89PNG\r\n\x1a\n")  # minimal PNG header

        # Build a minimal watchlist YAML (written before patching Path methods)
        wl_yaml = tmp_path / "test_watchlist.yaml"
        wl_yaml.write_text(
            "name: Test TII75\ntickers:\n  - ticker: CPRT\n    name: Copart Inc.\n"
        )

        with (
            patch("wally.main.fetch_price_snapshot", return_value=fake_snap),
            patch("wally.main.fetch_valuation_snapshot", return_value=fake_val_snap),
            patch("wally.main.render_range_chart", return_value=range_png_path),
            patch("wally.main.render_value_vs_price_chart", return_value=(None, "No valuation config found yet for this ticker")),
            patch("wally.main._build_xlsx", side_effect=FileNotFoundError("no config")),
            patch("wally.main._build_valuation_workbook") as mock_build_wb,
            patch("wally.main._analyse_opportunity", return_value=None),
            patch("wally.main.send_email", return_value=True),
            patch("wally.main.load_email_settings"),
            patch("wally.main.build_run_context") as mock_ctx,
            patch("wally.main.write_json"),
            patch("wally.main._XLSX_BUILDER_AVAILABLE", True),
            patch("wally.main._VALUATION_WORKBOOK_AVAILABLE", True),
        ):
            # Use a write_text that does nothing for the dashboard JSON
            with patch("wally.main.Path") as mock_path_cls:
                # Make Path("docs/data/wally.json") a no-op but keep real tmp_path
                mock_dash = MagicMock()
                mock_dash.exists.return_value = False
                mock_dash.parent.mkdir.return_value = None
                mock_path_cls.side_effect = lambda *a, **kw: (
                    mock_dash if "docs/data" in str(a[0]) else Path(*a, **kw)
                )

                # Configure run context to use tmp_path as output root
                run_ctx = MagicMock()
                run_ctx.output_root = tmp_path
                run_ctx.run_dt.strftime.return_value = "260317"
                mock_ctx.return_value = run_ctx

                # No Drive upload during test
                os.environ.pop("GDRIVE_FOLDER_ID", None)

                from wally.main import _process_watchlist

                result = _process_watchlist(
                    str(wl_yaml),
                    force=True,
                    send_individual_email=False,
                    is_tii75=False,
                )

        # The range PNG must be in inline_images with the correct CID so the
        # email body renders a chart image instead of plain text.
        cid_expected = "chart_cprt"
        inline_cids = [cid for cid, _ in result.inline_images]
        assert cid_expected in inline_cids, (
            f"Expected '{cid_expected}' in inline_images {inline_cids}. "
            "The range-chart PNG must be registered inline so TII75 emails "
            "show a visual chart (not just text)."
        )

        # The xlsx must also be in attachments
        xlsx_paths = [p for p in result.attachments if Path(p).suffix == ".xlsx"]
        assert xlsx_paths, "Expected fallback xlsx to be in attachments"

        # build_valuation_workbook must have been called once
        mock_build_wb.assert_called_once()


# ---------------------------------------------------------------------------
# Universal ticker normalization tests
# ---------------------------------------------------------------------------

class TestTickerNormalization:
    """Verify _ticker_slug works for all market formats."""

    def test_asx_ticker_slug(self):
        from wally.value_chart_builder import _ticker_slug
        assert _ticker_slug("NHC.AX") == "nhc_ax"

    def test_us_ticker_slug(self):
        from wally.value_chart_builder import _ticker_slug
        assert _ticker_slug("POOL") == "pool"
        assert _ticker_slug("FICO") == "fico"
        assert _ticker_slug("MSFT") == "msft"

    def test_tokyo_ticker_slug(self):
        from wally.value_chart_builder import _ticker_slug
        assert _ticker_slug("2914.T") == "2914_t"

    def test_infer_exchange_asx(self):
        from wally.value_chart_builder import _infer_exchange
        assert _infer_exchange("NHC.AX") == "ASX"
        assert _infer_exchange("BHP.AX") == "ASX"

    def test_infer_exchange_us_plain(self):
        from wally.value_chart_builder import _infer_exchange
        assert _infer_exchange("POOL") == "NASDAQ"
        assert _infer_exchange("FICO") == "NASDAQ"
        assert _infer_exchange("MSFT") == "NASDAQ"

    def test_infer_exchange_tokyo(self):
        from wally.value_chart_builder import _infer_exchange
        assert _infer_exchange("2914.T") == "TSE"

    def test_infer_currency_aud(self):
        from wally.value_chart_builder import _infer_currency
        assert _infer_currency("ASX") == "AUD"

    def test_infer_currency_usd(self):
        from wally.value_chart_builder import _infer_currency
        assert _infer_currency("NASDAQ") == "USD"
        assert _infer_currency("NYSE") == "USD"

    def test_infer_currency_jpy(self):
        from wally.value_chart_builder import _infer_currency
        assert _infer_currency("TSE") == "JPY"


# ---------------------------------------------------------------------------
# Auto-create starter config tests
# ---------------------------------------------------------------------------

class TestAutoCreateStarterConfig:
    """Verify starter configs are auto-created when missing."""

    def test_auto_create_for_us_ticker(self, tmp_path):
        """load_config auto-creates a starter YAML for a US ticker with no config."""
        import yaml
        from wally.value_chart_builder import _ticker_slug

        yaml_path = tmp_path / f"{_ticker_slug('POOL')}.yaml"
        assert not yaml_path.exists()

        # Monkey-patch repo_root so auto-create writes to tmp_path
        with patch("wally.value_chart_builder.Path") as mock_path_cls:
            # Allow real Path() calls, but intercept the valuations/<slug>.yaml resolution
            _real_path = Path

            def _side_effect(*args, **kwargs):
                p = _real_path(*args, **kwargs)
                return p

            mock_path_cls.side_effect = _side_effect

            # Instead of complex patching, test the function directly
            from wally.value_chart_builder import _auto_create_starter_config
            cfg = _auto_create_starter_config("POOL", yaml_path)

        assert yaml_path.exists(), "Starter YAML should be created"
        assert cfg["ticker"] == "POOL"
        assert cfg["exchange"] == "NASDAQ"
        assert cfg["currency"] == "USD"
        assert cfg["earnings"] == []

    def test_auto_create_for_tokyo_ticker(self, tmp_path):
        from wally.value_chart_builder import _auto_create_starter_config
        yaml_path = tmp_path / "2914_t.yaml"
        cfg = _auto_create_starter_config("2914.T", yaml_path)
        assert cfg["exchange"] == "TSE"
        assert cfg["currency"] == "JPY"
        assert cfg["ticker"] == "2914.T"

    def test_auto_create_for_asx_ticker(self, tmp_path):
        from wally.value_chart_builder import _auto_create_starter_config
        yaml_path = tmp_path / "xyz_ax.yaml"
        cfg = _auto_create_starter_config("XYZ.AX", yaml_path)
        assert cfg["exchange"] == "ASX"
        assert cfg["currency"] == "AUD"


# ---------------------------------------------------------------------------
# validate_config no longer requires earnings
# ---------------------------------------------------------------------------

class TestValidateConfigRelaxed:
    """Verify that _validate_config allows missing/empty earnings."""

    def test_empty_earnings_allowed(self):
        from wally.value_chart_builder import _validate_config
        cfg = {
            "ticker": "POOL",
            "company_name": "Pool Corporation",
            "buy_multiple": 25,
            "rror": 0.03,
            "earnings": [],
        }
        _validate_config(cfg)  # Should NOT raise
        assert cfg["earnings"] == []

    def test_missing_earnings_key_defaults_to_empty(self):
        from wally.value_chart_builder import _validate_config
        cfg = {
            "ticker": "FICO",
            "company_name": "Fair Isaac Corporation",
            "buy_multiple": 30,
            "rror": 0.03,
        }
        _validate_config(cfg)
        assert cfg["earnings"] == []

    def test_exchange_inferred_when_missing(self):
        from wally.value_chart_builder import _validate_config
        cfg = {
            "ticker": "NHC.AX",
            "company_name": "New Hope Corporation",
            "buy_multiple": 7,
            "rror": 0.05,
        }
        _validate_config(cfg)
        assert cfg["exchange"] == "ASX"
        assert cfg["currency"] == "AUD"

    def test_missing_required_key_raises(self):
        from wally.value_chart_builder import _validate_config
        import pytest
        cfg = {
            "ticker": "POOL",
            "company_name": "Pool Corporation",
            # buy_multiple is missing
            "rror": 0.03,
        }
        with pytest.raises(ValueError, match="buy_multiple"):
            _validate_config(cfg)


# ---------------------------------------------------------------------------
# Canonical 5-sheet workbook output — with and without earnings
# ---------------------------------------------------------------------------

class TestCanonicalWorkbookSheets:
    """Verify the canonical 5-sheet workbook is always produced."""

    _REQUIRED_SHEETS = {"Settings", "EarningsData", "PriceData", "ValueChart", "FuturePrompt"}

    def _make_cfg(self, ticker: str, exchange: str, currency: str,
                  earnings: list | None = None) -> dict:
        return {
            "ticker": ticker,
            "company_name": f"{ticker} Corp",
            "exchange": exchange,
            "currency": currency,
            "buy_multiple": 15,
            "rror": 0.04,
            "earnings": earnings or [],
        }

    def test_all_5_sheets_present_with_earnings(self, tmp_path):
        """Full build with earnings produces all 5 required sheets."""
        import openpyxl
        from wally.value_chart_builder import _validate_config, _build_settings, _build_earnings, _build_price_data, _build_chart, _build_future_prompt, get_price_data
        from openpyxl import Workbook

        cfg = self._make_cfg("TESTUS", "NASDAQ", "USD", earnings=[
            {"date": "2022-01-15", "period": "H1 FY2022", "ttm_eps": 100, "ttm_div": 20, "notes": ""},
            {"date": "2022-07-15", "period": "FY2022",    "ttm_eps": 110, "ttm_div": 22, "notes": ""},
        ])
        _validate_config(cfg)

        # Use synthetic price data (no yfinance needed)
        import pandas as pd
        import numpy as np
        dates = pd.date_range("2020-01-01", periods=100, freq="W-FRI")
        price_df = pd.DataFrame({"Date": dates, "Close": np.random.uniform(10, 20, 100)})

        wb = Workbook()
        wb.remove(wb.active)
        _build_settings(wb, cfg)
        _build_earnings(wb, cfg)
        _build_price_data(wb, cfg, price_df)
        _build_chart(wb, cfg, wb["PriceData"])
        _build_future_prompt(wb, cfg)

        out = tmp_path / "test_full.xlsx"
        wb.save(out)

        wb2 = openpyxl.load_workbook(out)
        missing = self._REQUIRED_SHEETS - set(wb2.sheetnames)
        assert not missing, f"Missing sheets: {missing}"

    def test_all_5_sheets_present_empty_earnings(self, tmp_path):
        """Build with empty earnings still produces all 5 required sheets (DATA REQUIRED)."""
        import openpyxl
        from wally.value_chart_builder import _validate_config, _build_settings, _build_earnings, _build_price_data, _build_chart, _build_future_prompt
        from openpyxl import Workbook

        cfg = self._make_cfg("POOL", "NASDAQ", "USD", earnings=[])
        _validate_config(cfg)

        import pandas as pd
        import numpy as np
        dates = pd.date_range("2020-01-01", periods=50, freq="W-FRI")
        price_df = pd.DataFrame({"Date": dates, "Close": np.random.uniform(200, 400, 50)})

        wb = Workbook()
        wb.remove(wb.active)
        _build_settings(wb, cfg)
        _build_earnings(wb, cfg)
        _build_price_data(wb, cfg, price_df)
        _build_chart(wb, cfg, wb["PriceData"])
        _build_future_prompt(wb, cfg)

        out = tmp_path / "pool_empty.xlsx"
        wb.save(out)

        wb2 = openpyxl.load_workbook(out)
        missing = self._REQUIRED_SHEETS - set(wb2.sheetnames)
        assert not missing, f"Missing sheets: {missing}"

        # DATA REQUIRED warning should appear in EarningsData
        ws_earn = wb2["EarningsData"]
        cell_a3 = ws_earn["A3"].value or ""
        assert "DATA REQUIRED" in str(cell_a3), (
            "EarningsData sheet should contain DATA REQUIRED warning when earnings list is empty"
        )

    def test_settings_title_uses_exchange_label(self, tmp_path):
        """Settings sheet title shows exchange label, not hardcoded 'ASX'."""
        import openpyxl
        from wally.value_chart_builder import _validate_config, _build_settings
        from openpyxl import Workbook

        for ticker, exchange, expected_label in [
            ("POOL",   "NASDAQ", "NASDAQ: POOL"),
            ("NHC.AX", "ASX",    "ASX: NHC"),
            ("2914.T", "TSE",    "TSE: 2914"),
        ]:
            cfg = self._make_cfg(ticker, exchange, "USD")
            _validate_config(cfg)
            wb = Workbook()
            wb.remove(wb.active)
            _build_settings(wb, cfg)
            out = tmp_path / f"{ticker.replace('.','_')}_settings.xlsx"
            wb.save(out)
            wb2 = openpyxl.load_workbook(out)
            title = wb2["Settings"]["A1"].value or ""
            assert expected_label in title, (
                f"Expected '{expected_label}' in Settings title, got: {title!r}"
            )

    def test_earnings_data_title_uses_exchange_label(self, tmp_path):
        """EarningsData title uses exchange label, not hardcoded 'ASX'."""
        import openpyxl
        from wally.value_chart_builder import _validate_config, _build_earnings
        from openpyxl import Workbook

        cfg = self._make_cfg("2914.T", "TSE", "JPY", earnings=[
            {"date": "2022-03-31", "period": "FY2022", "ttm_eps": 300, "ttm_div": 50, "notes": ""},
        ])
        _validate_config(cfg)
        wb = Workbook()
        wb.remove(wb.active)
        _build_earnings(wb, cfg)
        out = tmp_path / "tokyo_earnings.xlsx"
        wb.save(out)
        wb2 = openpyxl.load_workbook(out)
        title = wb2["EarningsData"]["A1"].value or ""
        assert "TSE: 2914" in title, f"Expected 'TSE: 2914' in title, got: {title!r}"
        # Column headers should use JPY not AUD
        h3 = wb2["EarningsData"]["C2"].value or ""
        assert "JPY" in h3, f"Expected 'JPY' in EPS column header, got: {h3!r}"


# ---------------------------------------------------------------------------
# main.py: explicit ERROR log before fallback
# ---------------------------------------------------------------------------

class TestMainLogsErrorBeforeFallback:
    """Verify main.py logs the exact error reason before creating fallback."""

    def test_error_logged_when_full_build_fails(self, tmp_path, capsys):
        """When _build_xlsx raises any exception, main logs ERROR before fallback."""
        import os
        from unittest.mock import MagicMock, patch
        from wally.data_fetch import PriceSnapshot, ValuationSnapshot

        fake_snap = PriceSnapshot(
            ticker="POOL",
            company_name="Pool Corporation",
            current_price=208.0,
            low_52w=203.0,
            high_52w=353.0,
        )
        fake_val_snap = ValuationSnapshot(
            trailing_pe=25.0, forward_pe=22.0, ev_to_ebitda=18.0,
            price_to_sales=4.5, fcf_yield=0.04, dividend_yield=None,
        )
        range_png = tmp_path / "pool_range.png"
        range_png.write_bytes(b"\x89PNG\r\n\x1a\n")

        wl_yaml = tmp_path / "test_wl.yaml"
        wl_yaml.write_text("name: Test\ntickers:\n  - ticker: POOL\n    name: Pool Corp\n")

        error_reason = "synthetic build failure for test"

        with (
            patch("wally.main.fetch_price_snapshot", return_value=fake_snap),
            patch("wally.main.fetch_valuation_snapshot", return_value=fake_val_snap),
            patch("wally.main.render_range_chart", return_value=range_png),
            patch("wally.main.render_value_vs_price_chart", return_value=(None, "no config")),
            patch("wally.main._build_xlsx", side_effect=RuntimeError(error_reason)),
            patch("wally.main._build_valuation_workbook"),
            patch("wally.main._analyse_opportunity", return_value=None),
            patch("wally.main.send_email"),
            patch("wally.main.load_email_settings"),
            patch("wally.main.build_run_context") as mock_ctx,
            patch("wally.main.write_json"),
            patch("wally.main._XLSX_BUILDER_AVAILABLE", True),
            patch("wally.main._VALUATION_WORKBOOK_AVAILABLE", True),
        ):
            with patch("wally.main.Path") as mock_path_cls:
                _real_path = Path
                mock_dash = MagicMock()
                mock_dash.exists.return_value = False
                mock_dash.parent.mkdir.return_value = None
                mock_path_cls.side_effect = lambda *a, **kw: (
                    mock_dash if "docs/data" in str(a[0]) else _real_path(*a, **kw)
                )
                run_ctx = MagicMock()
                run_ctx.output_root = tmp_path
                run_ctx.run_dt.strftime.return_value = "260317"
                mock_ctx.return_value = run_ctx
                os.environ.pop("GDRIVE_FOLDER_ID", None)

                from wally.main import _process_watchlist
                _process_watchlist(str(wl_yaml), force=True, send_individual_email=False)

        captured = capsys.readouterr()
        assert "ERROR building full workbook for POOL" in captured.out, (
            "Expected '[wally] ERROR building full workbook for POOL:' in stdout.\n"
            f"Actual stdout:\n{captured.out}"
        )
        assert error_reason in captured.out, (
            f"Expected the error reason '{error_reason}' in log output"
        )

    def test_fallback_filename_contains_fallback_review(self, tmp_path, capsys):
        """Fallback workbook is logged with '_fallback_review' in the filename."""
        import os
        from unittest.mock import MagicMock, patch
        from wally.data_fetch import PriceSnapshot, ValuationSnapshot

        fake_snap = PriceSnapshot(
            ticker="FICO",
            company_name="Fair Isaac",
            current_price=1700.0,
            low_52w=1680.0,
            high_52w=2200.0,
        )
        fake_val_snap = ValuationSnapshot(
            trailing_pe=60.0, forward_pe=50.0, ev_to_ebitda=40.0,
            price_to_sales=20.0, fcf_yield=0.01, dividend_yield=None,
        )
        range_png = tmp_path / "fico_range.png"
        range_png.write_bytes(b"\x89PNG\r\n\x1a\n")

        wl_yaml = tmp_path / "test_wl_fico.yaml"
        wl_yaml.write_text("name: Test\ntickers:\n  - ticker: FICO\n    name: Fair Isaac\n")

        with (
            patch("wally.main.fetch_price_snapshot", return_value=fake_snap),
            patch("wally.main.fetch_valuation_snapshot", return_value=fake_val_snap),
            patch("wally.main.render_range_chart", return_value=range_png),
            patch("wally.main.render_value_vs_price_chart", return_value=(None, "no config")),
            patch("wally.main._build_xlsx", side_effect=RuntimeError("no price data")),
            patch("wally.main._build_valuation_workbook"),
            patch("wally.main._analyse_opportunity", return_value=None),
            patch("wally.main.send_email"),
            patch("wally.main.load_email_settings"),
            patch("wally.main.build_run_context") as mock_ctx,
            patch("wally.main.write_json"),
            patch("wally.main._XLSX_BUILDER_AVAILABLE", True),
            patch("wally.main._VALUATION_WORKBOOK_AVAILABLE", True),
        ):
            with patch("wally.main.Path") as mock_path_cls:
                _real_path = Path
                mock_dash = MagicMock()
                mock_dash.exists.return_value = False
                mock_dash.parent.mkdir.return_value = None
                mock_path_cls.side_effect = lambda *a, **kw: (
                    mock_dash if "docs/data" in str(a[0]) else _real_path(*a, **kw)
                )
                run_ctx = MagicMock()
                run_ctx.output_root = tmp_path
                run_ctx.run_dt.strftime.return_value = "260317"
                mock_ctx.return_value = run_ctx
                os.environ.pop("GDRIVE_FOLDER_ID", None)

                from wally.main import _process_watchlist
                _process_watchlist(str(wl_yaml), force=True, send_individual_email=False)

        captured = capsys.readouterr()
        assert "Fallback workbook created" in captured.out, (
            "Expected 'Fallback workbook created:' in log output.\n"
            f"Actual stdout:\n{captured.out}"
        )
        assert "fallback_review" in captured.out, (
            "Expected '_fallback_review' in the fallback filename log"
        )


# ---------------------------------------------------------------------------
# Live fundamentals fetch — earnings and dividend wiring
# ---------------------------------------------------------------------------

class TestLiveFundamentalsFetch:
    """Tests confirming live fundamentals are wired into workbook generation."""

    def _make_av_earnings_data(self, n_quarters: int = 8):
        """Return a fake EarningsData with *n_quarters* quarters of data."""
        from wally.alphavantage import AnnualEarning, EarningsData, QuarterlyEarning

        quarters = []
        for i in range(n_quarters):
            # Quarters are most-recent-first; fiscal dates step back 3 months each
            year = 2024 - i // 4
            month = 12 - (i % 4) * 3
            if month <= 0:
                month += 12
                year -= 1
            fiscal = f"{year}-{month:02d}-{28 if month != 2 else 28:02d}"
            reported = f"{year}-{month:02d}-25"
            quarters.append(
                QuarterlyEarning(
                    fiscal_date=fiscal,
                    reported_date=reported,
                    reported_eps=3.50 + i * 0.10,  # EPS in dollars
                    estimated_eps=3.40 + i * 0.10,
                    surprise=0.10,
                    surprise_pct=2.9,
                )
            )
        annual = [AnnualEarning(fiscal_date="2024-12-31", reported_eps=14.0)]
        return EarningsData(ticker="POOL", annual=annual, quarterly=quarters)

    def _make_div_series(self):
        """Return a fake yfinance dividends Series (4 × $1.10 = $4.40 TTM)."""
        import pandas as pd

        dates = pd.to_datetime(
            ["2024-09-15", "2024-06-15", "2024-03-15", "2023-12-15"]
        )
        return pd.Series([1.10, 1.10, 1.10, 1.10], index=dates)

    # ── build_workbook_earnings_history ──────────────────────────────────────

    def test_build_workbook_earnings_history_basic(self):
        """build_workbook_earnings_history converts AV earnings into workbook format."""
        from wally.alphavantage import build_workbook_earnings_history

        av_data = self._make_av_earnings_data(8)
        history = build_workbook_earnings_history(av_data)

        assert len(history) == 5, f"Expected 5 entries (8-3), got {len(history)}"
        first = history[0]
        assert "date" in first
        assert "period" in first
        assert "ttm_eps" in first
        assert "ttm_div" in first
        assert first["ttm_eps"] > 0, "TTM EPS should be positive"
        assert first["notes"] == "alphavantage"

    def test_build_workbook_earnings_history_sorted_oldest_first(self):
        """build_workbook_earnings_history returns entries oldest-first."""
        from wally.alphavantage import build_workbook_earnings_history

        av_data = self._make_av_earnings_data(8)
        history = build_workbook_earnings_history(av_data)

        dates = [e["date"] for e in history]
        assert dates == sorted(dates), "History must be sorted oldest-first"

    def test_build_workbook_earnings_history_with_dividends(self):
        """build_workbook_earnings_history computes TTM dividends from div series."""
        from wally.alphavantage import build_workbook_earnings_history

        av_data = self._make_av_earnings_data(8)
        div_series = self._make_div_series()
        history = build_workbook_earnings_history(av_data, div_series)

        # The most recent entry should have non-zero TTM div
        # (4 payments × $1.10 = $4.40 = 440 cents)
        last = history[-1]
        assert last["ttm_div"] > 0, "TTM Div should be populated from dividend series"

    def test_build_workbook_earnings_history_no_div_gives_zero_div(self):
        """Without dividend data, TTM Div should be 0 (not None)."""
        from wally.alphavantage import build_workbook_earnings_history

        av_data = self._make_av_earnings_data(8)
        history = build_workbook_earnings_history(av_data, div_series=None)

        for entry in history:
            assert entry["ttm_div"] == 0.0, (
                f"Expected 0 TTM Div when no dividends passed, got {entry['ttm_div']}"
            )

    def test_build_workbook_earnings_history_insufficient_quarters(self):
        """Returns empty list when fewer than 4 quarters of data are available."""
        from wally.alphavantage import build_workbook_earnings_history

        av_data = self._make_av_earnings_data(3)
        history = build_workbook_earnings_history(av_data)

        assert history == [], "Should return [] when < 4 quarters available"

    def test_build_workbook_earnings_history_ttm_eps_in_cents(self):
        """TTM EPS is returned in cents (eps_scale=100), not dollars."""
        from wally.alphavantage import build_workbook_earnings_history

        av_data = self._make_av_earnings_data(4)
        # 4 quarters of $3.50 EPS each → TTM = $14.00 → 1400 cents
        history = build_workbook_earnings_history(av_data, eps_scale=100.0)

        assert len(history) == 1
        # TTM EPS should be around 1400 cents (4 × ~3.5 dollars × 100)
        assert history[0]["ttm_eps"] > 100, (
            f"TTM EPS should be in cents (>100), got {history[0]['ttm_eps']}"
        )

    # ── _fetch_live_fundamentals ─────────────────────────────────────────────

    def test_fetch_live_fundamentals_calls_av_when_key_set(self, capsys):
        """_fetch_live_fundamentals calls Alpha Vantage when ALPHAVANTAGE_API_KEY is set."""
        from unittest.mock import MagicMock, patch

        av_data = self._make_av_earnings_data(8)
        fake_divs = self._make_div_series()

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", return_value=av_data) as mock_av,
            patch("yfinance.Ticker") as mock_yf,
        ):
            mock_yf.return_value.dividends = fake_divs

            from wally.value_chart_builder import _fetch_live_fundamentals

            history, earnings_src, div_src, warning = _fetch_live_fundamentals(
                "POOL", {"ticker": "POOL"}
            )

        mock_av.assert_called_once_with("POOL", "test-key")
        assert earnings_src == "alphavantage"
        assert len(history) > 0, "Should return earnings history when AV succeeds"
        assert warning == "", "No warning expected on successful fetch"

    def test_fetch_live_fundamentals_price_only_when_no_key(self, capsys):
        """_fetch_live_fundamentals returns PRICE_ONLY when no AV key is set."""
        from unittest.mock import patch

        import pandas as pd

        fake_divs = self._make_div_series()

        with (
            patch.dict("os.environ", {}, clear=True),
            patch("yfinance.Ticker") as mock_yf,
        ):
            mock_yf.return_value.dividends = fake_divs

            from wally.value_chart_builder import _fetch_live_fundamentals

            history, earnings_src, div_src, warning = _fetch_live_fundamentals(
                "POOL", {"ticker": "POOL"}
            )

        assert history == [], "No earnings when AV key missing"
        assert "ALPHAVANTAGE_API_KEY" in warning, (
            "Warning should mention missing API key"
        )

        captured = capsys.readouterr()
        assert "PRICE_ONLY_BUILD" in captured.out

    def test_fetch_live_fundamentals_logs_sources(self, capsys):
        """_fetch_live_fundamentals logs price_source, earnings_source, dividends_source."""
        from unittest.mock import patch

        av_data = self._make_av_earnings_data(8)
        fake_divs = self._make_div_series()

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", return_value=av_data),
            patch("yfinance.Ticker") as mock_yf,
        ):
            mock_yf.return_value.dividends = fake_divs

            from wally.value_chart_builder import _fetch_live_fundamentals

            _fetch_live_fundamentals("POOL", {"ticker": "POOL"})

        captured = capsys.readouterr()
        assert "price_source=yfinance" in captured.out
        assert "earnings_source=alphavantage" in captured.out
        assert "earnings_records=" in captured.out
        assert "dividend_records=" in captured.out
        assert "FULL_BUILD" in captured.out

    def test_fetch_live_fundamentals_logs_price_only_on_av_failure(self, capsys):
        """_fetch_live_fundamentals logs PRICE_ONLY_BUILD when AV call fails."""
        from unittest.mock import patch

        import pandas as pd

        fake_divs = self._make_div_series()

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", side_effect=RuntimeError("API error")),
            patch("yfinance.Ticker") as mock_yf,
        ):
            mock_yf.return_value.dividends = fake_divs

            from wally.value_chart_builder import _fetch_live_fundamentals

            history, _, _, warning = _fetch_live_fundamentals("POOL", {"ticker": "POOL"})

        assert history == []
        assert "WARNING" in warning
        captured = capsys.readouterr()
        assert "PRICE_ONLY_BUILD" in captured.out

    # ── build_value_chart integration ────────────────────────────────────────

    def _base_cfg(self, ticker: str = "POOL") -> dict:
        """Return a minimal in-memory config for integration tests (no file I/O)."""
        return {
            "ticker": ticker,
            "company_name": f"{ticker} Corporation",
            "exchange": "NASDAQ",
            "currency": "USD",
            "buy_multiple": 25,
            "rror": 0.04,
            "earnings": [],
        }

    def test_build_value_chart_calls_fetch_when_earnings_empty(self, tmp_path):
        """build_value_chart calls _fetch_live_fundamentals when config has empty earnings."""
        from unittest.mock import patch

        import numpy as np
        import pandas as pd

        av_data = self._make_av_earnings_data(8)
        fake_divs = self._make_div_series()

        fake_price_df = pd.DataFrame(
            {
                "Date": pd.date_range("2020-01-01", periods=50, freq="W-FRI"),
                "Close": np.random.uniform(200, 400, 50),
            }
        )

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", return_value=av_data) as mock_av,
            patch("yfinance.Ticker") as mock_yf,
            patch("wally.value_chart_builder.load_config", return_value=self._base_cfg("POOL")),
        ):
            mock_yf.return_value.dividends = fake_divs
            mock_yf.return_value.history.return_value = fake_price_df.set_index("Date")
            # Provide price data via monkeypatch to avoid live yfinance call for prices
            with patch(
                "wally.value_chart_builder.get_price_data",
                return_value=fake_price_df,
            ):
                from wally.value_chart_builder import build_value_chart

                out_path = str(tmp_path / "pool_test.xlsx")
                build_value_chart("POOL", output_path=out_path)

        mock_av.assert_called_once_with("POOL", "test-key")

    def test_pool_ttm_eps_populated_in_workbook(self, tmp_path):
        """POOL workbook has TTM EPS values in PriceData when AV fetch succeeds."""
        import openpyxl
        from unittest.mock import patch

        import numpy as np
        import pandas as pd

        av_data = self._make_av_earnings_data(8)
        fake_divs = self._make_div_series()

        fake_price_df = pd.DataFrame(
            {
                "Date": pd.date_range("2022-01-01", periods=150, freq="W-FRI"),
                "Close": np.random.uniform(200, 400, 150),
            }
        )

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", return_value=av_data),
            patch("yfinance.Ticker") as mock_yf,
            patch("wally.value_chart_builder.load_config", return_value=self._base_cfg("POOL")),
            patch(
                "wally.value_chart_builder.get_price_data",
                return_value=fake_price_df,
            ),
        ):
            mock_yf.return_value.dividends = fake_divs

            from wally.value_chart_builder import build_value_chart

            out_path = str(tmp_path / "pool_full.xlsx")
            build_value_chart("POOL", output_path=out_path)

        wb = openpyxl.load_workbook(out_path)
        ws = wb["PriceData"]

        # Check that TTM EPS column (col 3) has at least some non-None values
        eps_values = [ws.cell(row=r, column=3).value for r in range(3, ws.max_row + 1)]
        non_null = [v for v in eps_values if v is not None]
        assert non_null, (
            "PriceData TTM EPS column should be populated when AV data is fetched"
        )

        # And PE Ratio column should also have values
        pe_col = 8  # with no sell_multiple: Date(1) Price(2) EPS(3) VBuy(4) Div(5) DivRRoR(6) PE(7) PEsmooth(8)
        pe_values = [ws.cell(row=r, column=7).value for r in range(3, ws.max_row + 1)]
        non_null_pe = [v for v in pe_values if v is not None]
        assert non_null_pe, "PE Ratio column should be populated when TTM EPS is available"

    def test_dividend_paying_stock_populates_div_rror(self, tmp_path):
        """A stock with dividend history gets non-zero Div/RRoR in PriceData."""
        import openpyxl
        from unittest.mock import patch

        import numpy as np
        import pandas as pd

        av_data = self._make_av_earnings_data(8)
        # 4 quarterly dividends of $1.10
        fake_divs = self._make_div_series()

        fake_price_df = pd.DataFrame(
            {
                "Date": pd.date_range("2022-01-01", periods=150, freq="W-FRI"),
                "Close": np.random.uniform(200, 400, 150),
            }
        )

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", return_value=av_data),
            patch("yfinance.Ticker") as mock_yf,
            patch("wally.value_chart_builder.load_config", return_value=self._base_cfg("POOL")),
            patch(
                "wally.value_chart_builder.get_price_data",
                return_value=fake_price_df,
            ),
        ):
            mock_yf.return_value.dividends = fake_divs

            from wally.value_chart_builder import build_value_chart

            out_path = str(tmp_path / "pool_div.xlsx")
            build_value_chart("POOL", output_path=out_path)

        wb = openpyxl.load_workbook(out_path)

        # EarningsData TTM Div column (col 6) should have non-zero values
        ws_earn = wb["EarningsData"]
        div_vals = [
            ws_earn.cell(row=r, column=6).value
            for r in range(3, ws_earn.max_row + 1)
        ]
        non_zero_div = [v for v in div_vals if v is not None and v != 0]
        assert non_zero_div, (
            "EarningsData TTM Div should be non-zero for a dividend-paying stock"
        )

    def test_non_dividend_stock_still_populates_eps_value_lines(self, tmp_path):
        """A non-dividend stock still populates EPS-based Value Buy in PriceData."""
        import openpyxl
        from unittest.mock import patch

        import numpy as np
        import pandas as pd

        av_data = self._make_av_earnings_data(8)
        # No dividends
        empty_divs = pd.Series([], dtype=float)

        fake_price_df = pd.DataFrame(
            {
                "Date": pd.date_range("2022-01-01", periods=150, freq="W-FRI"),
                "Close": np.random.uniform(200, 400, 150),
            }
        )

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch("wally.alphavantage.fetch_earnings", return_value=av_data),
            patch("yfinance.Ticker") as mock_yf,
            patch("wally.value_chart_builder.load_config", return_value=self._base_cfg("FICO")),
            patch(
                "wally.value_chart_builder.get_price_data",
                return_value=fake_price_df,
            ),
        ):
            mock_yf.return_value.dividends = empty_divs

            from wally.value_chart_builder import build_value_chart

            out_path = str(tmp_path / "fico_no_div.xlsx")
            build_value_chart("FICO", output_path=out_path)

        wb = openpyxl.load_workbook(out_path)
        ws = wb["PriceData"]

        # Value Buy col (col 4) should be populated even without dividends
        vbuy_vals = [ws.cell(row=r, column=4).value for r in range(3, ws.max_row + 1)]
        non_null = [v for v in vbuy_vals if v is not None]
        assert non_null, (
            "PriceData Value Buy should be populated even for a non-dividend stock"
        )

    def test_blank_earnings_not_silent_when_fetch_fails(self, tmp_path):
        """When fundamentals fetch fails, EarningsData shows a visible WARNING message."""
        import openpyxl
        from unittest.mock import patch

        import numpy as np
        import pandas as pd

        fake_price_df = pd.DataFrame(
            {
                "Date": pd.date_range("2020-01-01", periods=20, freq="W-FRI"),
                "Close": np.random.uniform(200, 400, 20),
            }
        )

        with (
            patch.dict("os.environ", {"ALPHAVANTAGE_API_KEY": "test-key"}),
            patch(
                "wally.alphavantage.fetch_earnings",
                side_effect=RuntimeError("rate limited"),
            ),
            patch("yfinance.Ticker") as mock_yf,
            patch("wally.value_chart_builder.load_config", return_value=self._base_cfg("POOL")),
            patch(
                "wally.value_chart_builder.get_price_data",
                return_value=fake_price_df,
            ),
        ):
            mock_yf.return_value.dividends = pd.Series([], dtype=float)

            from wally.value_chart_builder import build_value_chart

            out_path = str(tmp_path / "pool_fail.xlsx")
            build_value_chart("POOL", output_path=out_path)

        wb = openpyxl.load_workbook(out_path)
        ws_earn = wb["EarningsData"]

        # Cell A3 must contain a visible warning — not be blank
        a3 = str(ws_earn["A3"].value or "")
        assert a3.strip() != "", "EarningsData A3 must not be blank when fetch fails"
        assert "WARNING" in a3 or "DATA REQUIRED" in a3 or "⚠" in a3, (
            f"EarningsData should contain a visible warning, got: {a3!r}"
        )
