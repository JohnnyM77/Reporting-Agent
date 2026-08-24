"""Optional transaction ledger.

Everything in Theo works without this module producing anything. If
``data/portfolio.xlsx`` is absent — which it will be until the file is copied
across from the PC — slides render with the judgement and no figures, and the
site shows dashes in the IRR columns. Nothing raises.

When the file *is* present, every BUY is rebuilt as a separate **Decision**.
That is the whole point: three purchases of the same company on three dates at
three prices are three decisions with three different answers, and averaging
them into one line item destroys the only information worth having. A story
told from memory compresses a scaling-in into a single act of conviction; the
ledger does not.

Reconstruction rules
--------------------
* BUY opens a lot.
* DIVIDEND is allocated pro-rata across the lots open on the payment date,
  weighted by shares held.
* SELL consumes lots FIFO; each lot books its own share of the proceeds on
  that date.
* SPLIT scales every open lot.
* Shares still held are valued at the summary table's implied unit price
  (Value AUD / Shares), so the terminal cash flow matches the portfolio's own
  mark rather than a second price source.

IRR is then a plain XIRR over that lot's own cash flows, solved by bisection.
Lots with a cost under $100 carrying more than 100 shares are demerger scrip:
they arrived, they were not bought, and an IRR on a near-zero cost base is a
meaningless number. They are flagged and given none.
"""

from __future__ import annotations

import dataclasses
import datetime as dt
import json
import re
from pathlib import Path
from typing import Any, Iterable

DEFAULT_LEDGER_PATH = Path("data/portfolio.xlsx")

# Two shapes of workbook are understood, and both live in the repo rather than
# on someone's desktop. The transaction export is the original format: a
# summary table followed by per-ticker "XYZ Transactions" blocks, from which
# decisions are reconstructed. The decision workbook is already reconstructed —
# one row per buy, with the cash flows on their own sheet — so it is read
# directly rather than rebuilt. Searched in order; the first that exists wins.
LEDGER_SEARCH_PATH = (
    DEFAULT_LEDGER_PATH,
    Path("data/JM_Decision_Level_IRR.xlsx"),
)

# The money-free export of the decision workbook, written by
# scripts/export_ledger.py. This is the file that is safe to commit to a public
# repository, and the one Theo reads in CI. Cash flows in it are normalised so
# each buy is -1.0, which leaves IRR and multiple untouched — see the script.
DECISIONS_JSON = Path("data/decisions.json")

DECISION_SHEET = "Decision IRR"
CASHFLOW_SHEET = "Cash Flows"

# Rows on the decision sheet whose first cell is not a decision id are notes,
# blank spacers, or the PORTFOLIO total line. None of them is a decision.
_DECISION_ID_RE = re.compile(r"^D\d+$")


def default_ledger_path() -> Path:
    """The first workbook that actually exists, or the canonical default."""
    for candidate in LEDGER_SEARCH_PATH:
        if candidate.is_file():
            return candidate
    return DEFAULT_LEDGER_PATH

# A lot that cost less than this, while carrying more than this many shares,
# did not get bought — it fell out of a demerger.
SCRIP_MAX_COST = 100.0
SCRIP_MIN_SHARES = 100.0

_BLOCK_RE = re.compile(r"^\s*(?:ASX[:\s]+)?([A-Z0-9]{2,6})\s+Transactions\b", re.IGNORECASE)
_NUM_RE = re.compile(r"-?[\d,]*\.?\d+")
_CCY_RE = re.compile(r"\b([A-Z]{3})\b")

TXN_TYPES = ("BUY", "SELL", "DIVIDEND", "SPLIT")


# --------------------------------------------------------------------------
# Cell coercion — spreadsheets are not a data format, they are a rumour
# --------------------------------------------------------------------------


def _num(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    match = _NUM_RE.search(text.replace("$", ""))
    if not match:
        return None
    try:
        out = float(match.group(0).replace(",", ""))
    except ValueError:
        return None
    return -out if negative else out


def _ccy(value: Any) -> str:
    if value is None:
        return ""
    match = _CCY_RE.search(str(value).upper())
    return match.group(1) if match else ""


def _date(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d %b %Y", "%d %B %Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


# --------------------------------------------------------------------------
# XIRR
# --------------------------------------------------------------------------


def npv(rate: float, flows: list[tuple[dt.date, float]]) -> float:
    if not flows:
        return 0.0
    start = min(f[0] for f in flows)
    total = 0.0
    for when, amount in flows:
        years = (when - start).days / 365.25
        total += amount / ((1.0 + rate) ** years)
    return total


def xirr(flows: list[tuple[dt.date, float]], lo: float = -0.9999, hi: float = 10.0) -> float | None:
    """Solve NPV(rate) == 0 by bisection. None when there is no sign change.

    Bisection rather than Newton because the function is well behaved on a
    bracketed interval and cannot run away — a wrong IRR that looks plausible
    is worse than no IRR.
    """
    if len(flows) < 2:
        return None
    if not any(a < 0 for _, a in flows) or not any(a > 0 for _, a in flows):
        return None

    f_lo, f_hi = npv(lo, flows), npv(hi, flows)
    if f_lo * f_hi > 0:
        # Widen once before giving up.
        hi = 100.0
        f_hi = npv(hi, flows)
        if f_lo * f_hi > 0:
            return None

    for _ in range(200):
        mid = (lo + hi) / 2.0
        f_mid = npv(mid, flows)
        if abs(f_mid) < 1e-9 or (hi - lo) < 1e-10:
            return mid
        if f_lo * f_mid <= 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid
    return (lo + hi) / 2.0


# --------------------------------------------------------------------------
# Model
# --------------------------------------------------------------------------


@dataclasses.dataclass
class Transaction:
    ticker: str
    kind: str
    date: dt.date | None
    shares: float | None
    price: float | None
    cash_flow: float | None


@dataclasses.dataclass
class Decision:
    """One BUY, followed to its conclusion."""

    ticker: str
    index: int
    date: dt.date | None
    shares: float = 0.0
    price: float = 0.0
    cost: float = 0.0
    shares_remaining: float = 0.0
    dividends: float = 0.0
    proceeds: float = 0.0
    value_now: float = 0.0
    is_scrip: bool = False
    flows: list[tuple[dt.date, float]] = dataclasses.field(default_factory=list)
    as_at: dt.date | None = None
    # Share of total capital ever committed, 0..1. Populated by the money-free
    # export, where it replaces the dollar cost as the way to compare sizing.
    weight: float = 0.0
    # True when the flows are normalised to a buy of -1.0, which makes every
    # dollar-denominated field meaningless. Callers show dashes instead.
    normalised: bool = False

    @property
    def label(self) -> str:
        return f"{self.ticker} #{self.index}"

    @property
    def total_in(self) -> float:
        return self.dividends + self.proceeds + self.value_now

    @property
    def multiple(self) -> float | None:
        if self.cost <= 0:
            return None
        return self.total_in / self.cost

    @property
    def years(self) -> float | None:
        if not self.date or not self.as_at:
            return None
        return (self.as_at - self.date).days / 365.25

    @property
    def irr(self) -> float | None:
        if self.is_scrip:
            return None
        return xirr(self.flows)

    @property
    def open(self) -> bool:
        return self.shares_remaining > 1e-9


@dataclasses.dataclass
class Holding:
    ticker: str
    shares: float = 0.0
    price: float = 0.0
    value: float = 0.0
    cost_basis: float = 0.0
    unrealized: float = 0.0
    realized: float = 0.0
    dividends: float = 0.0
    currency_gains: float = 0.0
    total_gains: float = 0.0
    currency: str = "AUD"
    summary_irr: float | None = None
    decisions: list[Decision] = dataclasses.field(default_factory=list)
    weight: float = 0.0
    normalised: bool = False

    @property
    def unit_value(self) -> float:
        if self.shares > 0 and self.value:
            return self.value / self.shares
        return self.price or 0.0

    @property
    def irr(self) -> float | None:
        """Capital-weighted IRR across every decision in the name."""
        flows = [f for d in self.decisions if not d.is_scrip for f in d.flows]
        return xirr(sorted(flows, key=lambda f: f[0])) if flows else self.summary_irr

    @property
    def multiple(self) -> float | None:
        cost = sum(d.cost for d in self.decisions if not d.is_scrip)
        if cost <= 0:
            return None
        return sum(d.total_in for d in self.decisions if not d.is_scrip) / cost

    @property
    def years(self) -> float | None:
        spans = [d.years for d in self.decisions if d.years is not None]
        return max(spans) if spans else None


@dataclasses.dataclass
class Ledger:
    holdings: dict[str, Holding] = dataclasses.field(default_factory=dict)
    as_at: dt.date | None = None
    path: Path | None = None
    warnings: list[str] = dataclasses.field(default_factory=list)
    # True when loaded from the money-free export: IRR, multiple and years are
    # exact, every dollar figure is absent by construction.
    normalised: bool = False

    def __bool__(self) -> bool:
        return bool(self.holdings)

    def get(self, ticker: str) -> Holding | None:
        return self.holdings.get(str(ticker).upper())

    def decisions(self, ticker: str) -> list[Decision]:
        holding = self.get(ticker)
        return holding.decisions if holding else []

    def by_capital(self) -> list[Holding]:
        key = (lambda h: h.weight) if self.normalised else (lambda h: h.value or 0.0)
        return sorted(self.holdings.values(), key=key, reverse=True)

    def tickers(self) -> list[str]:
        return sorted(self.holdings)


# The empty ledger. Callers never branch on None.
EMPTY = Ledger()


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------


def _open_workbook(path: Path) -> Any | None:
    """Open the workbook, or None when openpyxl is not installed."""
    try:
        import openpyxl  # noqa: PLC0415 — optional dependency, by design
    except ImportError:
        return None
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _rows_from_sheet(path: Path) -> list[list[Any]] | None:
    try:
        import openpyxl  # noqa: PLC0415 — optional dependency, by design
    except ImportError:
        return None
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = None
    for name in workbook.sheetnames:
        if "transaction" in name.lower():
            sheet = workbook[name]
            break
    if sheet is None:
        sheet = workbook[workbook.sheetnames[0]]
    return [list(r) for r in sheet.iter_rows(values_only=True)]


def _looks_like_header(row: list[Any]) -> bool:
    cells = {_str(c).lower() for c in row if c is not None}
    return "type" in cells and "date" in cells


def _parse_summary(rows: list[list[Any]]) -> tuple[dict[str, Holding], int]:
    """Read the summary table at the top. Returns (holdings, first row after)."""
    holdings: dict[str, Holding] = {}
    header_idx = None
    columns: dict[str, int] = {}

    for i, row in enumerate(rows[:60]):
        labels = [_str(c).lower() for c in row]
        if "ticker" in labels and any("value" in c for c in labels):
            header_idx = i
            for j, label in enumerate(labels):
                if not label:
                    continue
                if label == "ticker":
                    columns["ticker"] = j
                elif label == "shares":
                    columns["shares"] = j
                elif label == "price":
                    columns["price"] = j
                elif "value" in label:
                    columns["value"] = j
                elif "cost" in label:
                    columns["cost"] = j
                elif "unrealiz" in label or "unrealis" in label:
                    columns["unrealized"] = j
                elif "realiz" in label or "realis" in label:
                    columns["realized"] = j
                elif "dividend" in label:
                    columns["dividends"] = j
                elif "currency" in label:
                    columns["currency_gains"] = j
                elif "total gain" in label:
                    columns["total_gains"] = j
                elif "irr" in label:
                    columns["irr"] = j
            break

    if header_idx is None:
        return holdings, 0

    last = header_idx
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        raw_ticker = _str(row[columns["ticker"]]) if columns.get("ticker") is not None else ""
        if not raw_ticker:
            if any(_str(c) for c in row):
                continue
            break
        if _BLOCK_RE.match(raw_ticker):
            break
        ticker = raw_ticker.split(":")[-1].strip().upper()
        if not re.fullmatch(r"[A-Z0-9]{2,6}", ticker):
            continue

        def cell(key: str) -> Any:
            idx = columns.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        irr_raw = _num(cell("irr"))
        if irr_raw is not None and abs(irr_raw) > 1.5:
            irr_raw = irr_raw / 100.0
        holdings[ticker] = Holding(
            ticker=ticker,
            shares=_num(cell("shares")) or 0.0,
            price=_num(cell("price")) or 0.0,
            value=_num(cell("value")) or 0.0,
            cost_basis=_num(cell("cost")) or 0.0,
            unrealized=_num(cell("unrealized")) or 0.0,
            realized=_num(cell("realized")) or 0.0,
            dividends=_num(cell("dividends")) or 0.0,
            currency_gains=_num(cell("currency_gains")) or 0.0,
            total_gains=_num(cell("total_gains")) or 0.0,
            currency=_ccy(cell("price")) or "AUD",
            summary_irr=irr_raw,
        )
        last = i
    return holdings, last + 1


def _parse_blocks(rows: list[list[Any]], start: int) -> dict[str, list[Transaction]]:
    """Read the per-ticker transaction blocks below the summary."""
    blocks: dict[str, list[Transaction]] = {}
    ticker: str | None = None
    columns: dict[str, int] = {}

    for row in rows[start:]:
        first = ""
        for cell in row:
            if _str(cell):
                first = _str(cell)
                break

        match = _BLOCK_RE.match(first)
        if match:
            ticker = match.group(1).upper()
            blocks.setdefault(ticker, [])
            columns = {}
            continue

        if ticker is None:
            continue

        if _looks_like_header(row):
            columns = {}
            for j, cell in enumerate(row):
                label = _str(cell).lower()
                if not label:
                    continue
                if label.startswith("type"):
                    columns["type"] = j
                elif label.startswith("date"):
                    columns["date"] = j
                elif label.startswith("share") or label.startswith("unit") or label.startswith("qty"):
                    columns["shares"] = j
                elif label.startswith("price"):
                    columns["price"] = j
                elif "cash" in label or "amount" in label or "value" in label:
                    columns["cash"] = j
            continue

        if not columns:
            continue

        kind = _str(row[columns["type"]]).upper() if columns.get("type") is not None else ""
        if kind not in TXN_TYPES:
            continue

        def cell(key: str) -> Any:
            idx = columns.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        blocks[ticker].append(
            Transaction(
                ticker=ticker,
                kind=kind,
                date=_date(cell("date")),
                shares=_num(cell("shares")),
                price=_num(cell("price")),
                cash_flow=_num(cell("cash")),
            )
        )

    return blocks


# --------------------------------------------------------------------------
# Decision reconstruction
# --------------------------------------------------------------------------


def _build_decisions(
    ticker: str,
    transactions: list[Transaction],
    unit_value: float,
    as_at: dt.date,
    warnings: list[str],
) -> list[Decision]:
    ordered = sorted(
        [t for t in transactions if t.date],
        key=lambda t: (t.date, TXN_TYPES.index(t.kind) if t.kind in TXN_TYPES else 9),
    )
    lots: list[Decision] = []

    for txn in ordered:
        if txn.kind == "BUY":
            shares = txn.shares or 0.0
            price = txn.price
            cost = abs(txn.cash_flow) if txn.cash_flow else None
            if cost is None:
                cost = shares * (price or 0.0)
            if price is None and shares:
                price = cost / shares
            lot = Decision(
                ticker=ticker,
                index=len(lots) + 1,
                date=txn.date,
                shares=shares,
                price=price or 0.0,
                cost=cost or 0.0,
                shares_remaining=shares,
                as_at=as_at,
            )
            lot.is_scrip = lot.cost < SCRIP_MAX_COST and lot.shares > SCRIP_MIN_SHARES
            lot.flows.append((txn.date, -lot.cost))
            lots.append(lot)

        elif txn.kind == "DIVIDEND":
            amount = txn.cash_flow
            if amount is None and txn.shares and txn.price:
                amount = txn.shares * txn.price
            if not amount:
                continue
            amount = abs(amount)
            open_lots = [lot for lot in lots if lot.shares_remaining > 1e-9]
            base = sum(lot.shares_remaining for lot in open_lots)
            if base <= 0:
                continue
            for lot in open_lots:
                share = amount * (lot.shares_remaining / base)
                lot.dividends += share
                lot.flows.append((txn.date, share))

        elif txn.kind == "SELL":
            to_sell = abs(txn.shares or 0.0)
            price = txn.price
            if price is None and txn.cash_flow and to_sell:
                price = abs(txn.cash_flow) / to_sell
            price = price or 0.0
            for lot in lots:
                if to_sell <= 1e-9:
                    break
                if lot.shares_remaining <= 1e-9:
                    continue
                taken = min(lot.shares_remaining, to_sell)
                lot.shares_remaining -= taken
                to_sell -= taken
                cash = taken * price
                lot.proceeds += cash
                lot.flows.append((txn.date, cash))
            if to_sell > 1e-6:
                warnings.append(
                    f"{ticker}: sell on {txn.date} exceeded open lots by {to_sell:g} shares"
                )

        elif txn.kind == "SPLIT":
            open_lots = [lot for lot in lots if lot.shares_remaining > 1e-9]
            before = sum(lot.shares_remaining for lot in open_lots)
            raw = txn.shares if txn.shares is not None else txn.price
            if not raw or before <= 0:
                continue
            # The sheet writes either the post-split total or a bare ratio.
            ratio = (raw / before) if raw > before else raw
            if ratio <= 0:
                continue
            for lot in open_lots:
                lot.shares_remaining *= ratio
                lot.shares *= ratio
                if lot.shares:
                    lot.price = lot.cost / lot.shares

    for lot in lots:
        if lot.shares_remaining > 1e-9:
            lot.value_now = lot.shares_remaining * unit_value
            if lot.value_now:
                lot.flows.append((as_at, lot.value_now))
        lot.flows.sort(key=lambda f: f[0])

    return lots


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------


# --------------------------------------------------------------------------
# The decision workbook
# --------------------------------------------------------------------------


def _ticker_from_symbol(symbol: Any) -> str:
    """ASX:ABB -> ABB, LSE:RR. -> RR. Exchange prefixes are not part of the name."""
    text = str(symbol or "").strip().upper()
    if ":" in text:
        text = text.split(":", 1)[1]
    return text.rstrip(".").strip()


def _decision_flows(workbook: Any) -> tuple[dict[str, list[tuple[dt.date, float]]], dt.date | None]:
    """Cash flows per decision id, plus the date the terminal values were struck."""
    flows: dict[str, list[tuple[dt.date, float]]] = {}
    valued_at: dt.date | None = None
    if CASHFLOW_SHEET not in workbook.sheetnames:
        return flows, valued_at

    for row in workbook[CASHFLOW_SHEET].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        decision_id = str(row[0]).strip()
        if not _DECISION_ID_RE.match(decision_id):
            continue
        when = _date(row[2])
        amount = _num(row[4])
        if when is None or amount is None:
            continue
        flows.setdefault(decision_id, []).append((when, amount))
        # The terminal rows all carry the same valuation date. Trusting the
        # workbook's own mark date keeps `years` honest instead of silently
        # measuring to today against a price struck weeks ago.
        if str(row[3] or "").strip().upper() == "TERMINAL":
            valued_at = when if valued_at is None else max(valued_at, when)

    for entries in flows.values():
        entries.sort(key=lambda f: f[0])
    return flows, valued_at


def _load_decision_workbook(workbook: Any, ledger: Ledger, as_at: dt.date | None) -> bool:
    """Read an already-reconstructed decision workbook. True if anything loaded.

    The sibling parser rebuilds decisions from raw transactions. This one does
    not need to: the workbook has done it, and its own Reconciliation sheet
    shows the rebuilt IRRs matching. The cash flows are still re-solved here
    rather than reading the stored IRR column, so there is exactly one place in
    this repository where an IRR is computed.
    """
    if DECISION_SHEET not in workbook.sheetnames:
        return False

    flows, valued_at = _decision_flows(workbook)
    ledger.as_at = as_at or valued_at or ledger.as_at

    holdings: dict[str, Holding] = {}
    counts: dict[str, int] = {}

    for row in workbook[DECISION_SHEET].iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not _DECISION_ID_RE.match(str(row[0]).strip()):
            continue
        decision_id = str(row[0]).strip()
        ticker = _ticker_from_symbol(row[1])
        if not ticker:
            continue

        shares = _num(row[3]) or 0.0
        cost = _num(row[5]) or 0.0
        status = str(row[13] or "").strip().upper()
        holding = holdings.setdefault(ticker, Holding(ticker=ticker))
        counts[ticker] = counts.get(ticker, 0) + 1

        decision = Decision(
            ticker=ticker,
            index=counts[ticker],
            date=_date(row[2]),
            shares=shares,
            price=_num(row[4]) or 0.0,
            cost=cost,
            # No decision in this workbook is part-sold: a row either still
            # holds its shares or has none. If that ever changes, the sale
            # proceeds and the terminal value would both be non-zero and this
            # would need to split the parcel.
            shares_remaining=shares if status != "CLOSED" else 0.0,
            dividends=_num(row[7]) or 0.0,
            proceeds=_num(row[8]) or 0.0,
            value_now=_num(row[9]) or 0.0,
            is_scrip=cost < SCRIP_MAX_COST and shares > SCRIP_MIN_SHARES,
            flows=flows.get(decision_id, []),
            as_at=ledger.as_at,
        )
        holding.decisions.append(decision)

    for holding in holdings.values():
        holding.decisions.sort(key=lambda d: (d.date or dt.date.min, d.index))
        holding.shares = sum(d.shares_remaining for d in holding.decisions)
        holding.value = sum(d.value_now for d in holding.decisions)
        holding.cost_basis = sum(d.cost for d in holding.decisions)
        holding.dividends = sum(d.dividends for d in holding.decisions)
        holding.realized = sum(d.proceeds for d in holding.decisions)
        holding.unrealized = holding.value - sum(
            d.cost for d in holding.decisions if d.shares_remaining > 0
        )
        holding.total_gains = holding.unrealized + holding.realized + holding.dividends

    ledger.holdings = holdings
    return bool(holdings)


def _load_decisions_json(path: Path, ledger: Ledger, as_at: dt.date | None) -> bool:
    """Read the money-free export. True if anything loaded.

    Flows arrive normalised to a buy of -1.0, so every dollar figure is
    genuinely absent rather than hidden: cost, value and dividends stay None
    and render as dashes. IRR, multiple and years are exact, because none of
    them depends on the scale of the flows.
    """
    try:
        payload = json.loads(path.read_text())
    except (OSError, ValueError) as exc:
        ledger.warnings.append(f"could not read {path}: {exc}")
        return False

    valued_at = _date(payload.get("as_at"))
    ledger.as_at = as_at or valued_at or ledger.as_at

    holdings: dict[str, Holding] = {}
    for entry in payload.get("decisions", []):
        ticker = str(entry.get("ticker") or "").strip().upper()
        if not ticker:
            continue
        flows = []
        for raw_when, amount in entry.get("flows") or []:
            when = _date(raw_when)
            if when is not None:
                flows.append((when, float(amount)))
        # The file stores each decision's flows against a buy of -1.0, which is
        # right for that decision's own IRR and wrong the moment two decisions
        # are pooled: a $250 birthday present would count for as much as a
        # $263,000 cheque. Rescaling by the decision's share of total capital
        # fixes the pooling and, because IRR is scale-invariant, leaves the
        # decision's own IRR exactly where it was.
        weight = float(entry.get("weight") or 0.0)
        flows = [(when, amount * weight) for when, amount in flows]

        holding = holdings.setdefault(ticker, Holding(ticker=ticker))
        holding.decisions.append(
            Decision(
                ticker=ticker,
                index=int(entry.get("index") or len(holding.decisions) + 1),
                date=_date(entry.get("date")),
                price=float(entry.get("price") or 0.0),
                # Capital committed, expressed as a fraction of the portfolio
                # rather than in dollars.
                cost=weight,
                shares_remaining=weight if str(entry.get("status")).upper() != "CLOSED" else 0.0,
                value_now=sum(a for _, a in flows if a > 0),
                weight=weight,
                normalised=True,
                is_scrip=bool(entry.get("scrip")),
                flows=sorted(flows),
                as_at=ledger.as_at,
            )
        )

    for holding in holdings.values():
        holding.decisions.sort(key=lambda d: (d.date or dt.date.min, d.index))
        holding.normalised = True
        holding.weight = sum(d.weight for d in holding.decisions)

    ledger.holdings = holdings
    ledger.normalised = True
    return bool(holdings)


def available(path: str | Path | None = None) -> bool:
    if path:
        return Path(path).is_file()
    return default_ledger_path().is_file() or DECISIONS_JSON.is_file()


def load(path: str | Path | None = None, as_at: dt.date | None = None) -> Ledger:
    """Load the ledger. Returns an empty Ledger when the file is not there.

    Never raises for a missing file or a missing openpyxl — the ledger is
    optional and the rest of Theo is expected to run without it.
    """
    target = Path(path) if path else default_ledger_path()
    ledger = Ledger(path=target, as_at=as_at or dt.date.today())
    if target.suffix.lower() == ".json" and target.is_file():
        _load_decisions_json(target, ledger, as_at)
        return ledger
    if not target.is_file():
        # No private workbook — fall back to the money-free export, which is
        # the normal case in CI and on any machine that is not the one holding
        # the spreadsheet.
        if not path and DECISIONS_JSON.is_file():
            ledger.path = DECISIONS_JSON
            _load_decisions_json(DECISIONS_JSON, ledger, as_at)
        return ledger

    workbook = _open_workbook(target)
    if workbook is None:
        ledger.warnings.append(
            "openpyxl is not installed — ledger ignored (pip install openpyxl)"
        )
        return ledger
    if _load_decision_workbook(workbook, ledger, as_at):
        return ledger

    rows = _rows_from_sheet(target)
    if rows is None:
        ledger.warnings.append(
            "openpyxl is not installed — ledger ignored (pip install openpyxl)"
        )
        return ledger

    holdings, next_row = _parse_summary(rows)
    blocks = _parse_blocks(rows, next_row)

    for ticker, transactions in blocks.items():
        holding = holdings.get(ticker)
        if holding is None:
            holding = Holding(ticker=ticker)
            holdings[ticker] = holding
        holding.decisions = _build_decisions(
            ticker, transactions, holding.unit_value, ledger.as_at, ledger.warnings
        )

    ledger.holdings = holdings
    return ledger


def gaps(ledger: Ledger, covered: Iterable[str]) -> list[Holding]:
    """Holdings with real capital in them and no thesis written, by size."""
    have = {str(t).upper() for t in covered}
    missing = [h for t, h in ledger.holdings.items() if t not in have and (h.shares or 0) > 0]
    return sorted(missing, key=lambda h: h.value or 0.0, reverse=True)
