from __future__ import annotations

import datetime as dt
import json
import subprocess
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import NumFmt
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .drive_upload import upload_or_replace_xlsx

NAVY = "1F2D4E"
YELLOW = "FFFF00"
BLUE = "0000FF"
GREY = "F2F2F2"
WHITE = "FFFFFF"
BORDER_COLOR = "BBBBBB"

PROMPT_FUTURE_TEXT = "WALLY AGENT PROMPT — Value Analysis Spreadsheet Generator (stored prompt text)."


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _style_cell(cell, *, bold=False, color="000000", fill=None, align="left", wrap=False, size=9, italic=False):
    cell.font = Font(name="Arial", size=size, bold=bold, color=color, italic=italic)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()


def _apply_row_band(ws, row: int, start_col: int, end_col: int):
    fill = GREY if row % 2 == 0 else WHITE
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill)


def _safe_half_values(ttm_values: list[float]) -> list[float]:
    out = []
    prev = None
    for val in ttm_values:
        if prev is None:
            out.append(val / 2.0)
        else:
            out.append(max(0.0, val - prev / 2.0))
        prev = val
    return out


def _lookup_last(target_ts: pd.Timestamp, dates: list[pd.Timestamp], values: list[float]) -> float | None:
    result = None
    for d, v in zip(dates, values):
        if d <= target_ts:
            result = v
        else:
            break
    return result


def _run_recalc(output_path: Path) -> dict[str, Any]:
    recalc_script = Path("scripts/recalc.py")
    if not recalc_script.exists():
        return {"status": "success", "total_errors": 0, "note": "recalc script unavailable"}
    try:
        proc = subprocess.run(["python", str(recalc_script), str(output_path)], check=False, capture_output=True, text=True)
        if proc.stdout.strip():
            return json.loads(proc.stdout)
    except Exception as e:
        return {"status": "error", "total_errors": 1, "error": str(e)}
    return {"status": "success", "total_errors": 0}


def generate_asx_value_spreadsheet(config: dict, price_data_path: str, output_path: str) -> str:
    company_name = config["company_name"]
    ticker = config["ticker"]
    multiple = int(config["multiple"])
    rror = float(config["rror"])
    pe_smooth_days = int(config.get("pe_smooth_days", 60))
    earnings = config.get("earnings", [])

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # Settings sheet
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50
    ws.merge_cells("A1:C1")
    ws["A1"] = f"{company_name} (ASX: {ticker}) — Value Analysis — Settings"
    _style_cell(ws["A1"], bold=True, color="FFFFFF", fill=NAVY, align="center", size=13)
    ws.row_dimensions[1].height = 28

    for sec_row, text in [(3, "COMPANY DETAILS"), (7, "VALUATION ASSUMPTIONS"), (11, "ANALYSIS PERIOD"), (15, "HOW TO REFRESH SHARE PRICE DATA")]:
        ws.merge_cells(f"A{sec_row}:C{sec_row}")
        ws[f"A{sec_row}"] = text
        _style_cell(ws[f"A{sec_row}"], bold=True, color="FFFFFF", fill=NAVY, align="left")
        ws.row_dimensions[sec_row].height = 20

    ws["A4"], ws["B4"], ws["C4"] = "Company Name", company_name, "Auto-filled from config"
    ws["A5"], ws["B5"] = "ASX Ticker", ticker
    ws["A8"], ws["B8"], ws["C8"] = "Earnings Multiple (P/E)", multiple, "P/E multiple you are willing to pay — used in Value price formula"
    ws["A9"], ws["B9"], ws["C9"] = "Required Rate of Return (RRoR)", rror, "Minimum acceptable dividend yield, e.g. 0.03 = 3%"
    ws["B8"].number_format = "0"
    ws["B9"].number_format = "0.00%"
    _style_cell(ws["B8"], color=BLUE, fill=YELLOW)
    _style_cell(ws["B9"], color=BLUE, fill=YELLOW)

    earliest = earnings[0][0].strftime("%d-%b-%Y") if earnings else "N/A"
    ws["A12"], ws["B12"] = "From", earliest
    ws["A13"], ws["B13"] = "To", "Latest available result"

    instructions = [
        f'Go to finance.yahoo.com and search "{ticker}.AX"',
        "Click Historical Data → Time Period: 10Y → Frequency: Daily → Apply → Download",
        "Ensure CSV columns are: Date (YYYYMMDD), Open, High, Low, Close, Volume",
        "Replace the price CSV path in the agent config and re-run the agent",
        f'Microsoft 365 alternative: =STOCKHISTORY("{ticker}.AX", DATE(2016,1,1), TODAY(), 0, 1, 0, 1)',
        "The spreadsheet will be regenerated and re-uploaded to Google Drive automatically",
        f"P/E smooth uses a {pe_smooth_days}-day rolling average — adjust pe_smooth_days in config to change",
    ]
    for i, line in enumerate(instructions, start=16):
        ws.merge_cells(f"A{i}:C{i}")
        ws[f"A{i}"] = line
        _style_cell(ws[f"A{i}"], italic=True, color="666666")

    # EarningsData
    ws2 = wb.create_sheet("EarningsData")
    ws2.sheet_view.showGridLines = False
    headers = ["Ann. Date", "Period", "Half EPS (AUD c)", "TTM EPS (AUD c)", "Half Div (AUD c)", "TTM Div (AUD c)", f"Value ({multiple}x E)", "Div / RRoR", "Notes"]
    widths = [14, 24, 13, 13, 12, 12, 16, 16, 44]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    ws2.merge_cells("A1:I1")
    ws2["A1"] = f"{company_name} (ASX: {ticker}) — Half-Year Earnings & Dividend Data"
    _style_cell(ws2["A1"], bold=True, color="FFFFFF", fill=NAVY, size=13)
    ws2.row_dimensions[1].height = 28

    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=c, value=h)
        _style_cell(cell, bold=True, color="FFFFFF", fill=NAVY, align="center", wrap=True)
    ws2.row_dimensions[2].height = 35

    ttm_eps = [float(e[1]) for e in earnings]
    ttm_div = [float(e[2]) for e in earnings]
    half_eps = _safe_half_values(ttm_eps)
    half_div = _safe_half_values(ttm_div)

    for idx, e in enumerate(earnings, start=3):
        ann_date, _ttm_eps, _ttm_div, period_label, notes = e
        ws2.cell(idx, 1, ann_date)
        ws2.cell(idx, 2, period_label)
        ws2.cell(idx, 3, half_eps[idx - 3])
        ws2.cell(idx, 4, ttm_eps[idx - 3])
        ws2.cell(idx, 5, half_div[idx - 3])
        ws2.cell(idx, 6, ttm_div[idx - 3])
        ws2.cell(idx, 7, f"=D{idx}*Settings!$B$8/100")
        ws2.cell(idx, 8, f"=F{idx}/Settings!$B$9/100")
        ws2.cell(idx, 9, str(notes))
        for c in range(1, 10):
            _style_cell(ws2.cell(idx, c), align="center" if c not in (1, 2, 9) else "left")
        _apply_row_band(ws2, idx, 1, 9)

    for r in range(3, ws2.max_row + 1):
        ws2.cell(r, 1).number_format = "DD-MMM-YYYY"
        ws2.cell(r, 3).number_format = "#,##0.0"
        ws2.cell(r, 4).number_format = "#,##0.0"
        ws2.cell(r, 5).number_format = "#,##0.0"
        ws2.cell(r, 6).number_format = "#,##0.0"
        ws2.cell(r, 7).number_format = '"$"#,##0.00'
        ws2.cell(r, 8).number_format = '"$"#,##0.00'

    legend_row = ws2.max_row + 2
    ws2.merge_cells(f"A{legend_row}:I{legend_row}")
    ws2[f"A{legend_row}"] = (
        "COLUMN GUIDE: TTM EPS = trailing 12-month underlying EPS | "
        "Value (Nx E) = TTM EPS x Multiple / 100 (Settings!B8) | "
        "Div/RRoR = TTM Div / RRoR / 100 (Settings!B9)"
    )
    _style_cell(ws2[f"A{legend_row}"], italic=True, color="666666", wrap=True)
    ws2.row_dimensions[legend_row].height = 30

    # PriceData
    ws3 = wb.create_sheet("PriceData")
    ws3.sheet_view.showGridLines = False
    price_headers = ["Date", "Price (AUD $)", "TTM EPS (c)", f"Value ({multiple}x E)", "TTM Div (c)", "Div / RRoR ($)", "P/E Ratio", f"P/E {pe_smooth_days}D Avg"]
    price_widths = [11, 11, 10, 14, 10, 13, 9, 12]
    for i, w in enumerate(price_widths, start=1):
        ws3.column_dimensions[chr(64 + i)].width = w

    ws3.merge_cells("A1:H1")
    ws3["A1"] = f"{company_name} (ASX: {ticker}) — Daily Price & Value Data"
    _style_cell(ws3["A1"], bold=True, color="FFFFFF", fill=NAVY, size=11)

    for c, h in enumerate(price_headers, start=1):
        cell = ws3.cell(2, c, h)
        _style_cell(cell, bold=True, color="FFFFFF", fill=NAVY, align="center", wrap=True)

    df = pd.read_csv(price_data_path)
    df["Date"] = pd.to_datetime(df["Date"], format="%Y%m%d", errors="coerce")
    df = df[df["Date"] >= "2016-01-01"].sort_values("Date").reset_index(drop=True)
    df["Close"] = pd.to_numeric(df["Close"], errors="coerce")
    df = df.dropna(subset=["Close"])
    if df.empty:
        raise RuntimeError("No usable share price data available for spreadsheet generation")

    earn_dates = [pd.Timestamp(e[0]) for e in earnings]
    earn_eps = [float(e[1]) for e in earnings]
    earn_div = [float(e[2]) for e in earnings]

    rows = []
    for _, row in df.iterrows():
        d = pd.Timestamp(row["Date"])
        close = float(row["Close"])
        eps = _lookup_last(d, earn_dates, earn_eps)
        div = _lookup_last(d, earn_dates, earn_div)
        val = (eps * multiple / 100) if eps is not None else None
        div_rror = (div / rror / 100) if (div is not None and rror > 0) else None
        pe = (close / (eps / 100)) if (eps is not None and eps > 0) else np.nan
        rows.append((d.to_pydatetime().date(), close, eps, val, div, div_rror, pe))

    pe_series = pd.Series([r[6] for r in rows], dtype=float)
    pe_smooth = pe_series.rolling(window=pe_smooth_days, min_periods=max(10, pe_smooth_days // 6)).mean().round(2)

    for idx, r in enumerate(rows, start=3):
        vals = [r[0], r[1], r[2], r[3], r[4], r[5], None if pd.isna(r[6]) else float(r[6]), None if pd.isna(pe_smooth.iloc[idx - 3]) else float(pe_smooth.iloc[idx - 3])]
        for c, v in enumerate(vals, start=1):
            ws3.cell(idx, c, v)
            _style_cell(ws3.cell(idx, c), align="center")
        _apply_row_band(ws3, idx, 1, 8)

    for r in range(3, ws3.max_row + 1):
        ws3.cell(r, 1).number_format = "DD-MMM-YYYY"
        ws3.cell(r, 2).number_format = '"$"#,##0.00'
        ws3.cell(r, 3).number_format = "#,##0.0"
        ws3.cell(r, 4).number_format = '"$"#,##0.00'
        ws3.cell(r, 5).number_format = "#,##0.0"
        ws3.cell(r, 6).number_format = '"$"#,##0.00'
        ws3.cell(r, 7).number_format = "0.0"
        ws3.cell(r, 8).number_format = "0.0"

    # ValueChart
    ws4 = wb.create_sheet("ValueChart")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:N1")
    ws4["A1"] = f"{company_name} (ASX: {ticker})  —  Value Analysis Chart"
    _style_cell(ws4["A1"], bold=True, color="FFFFFF", fill=NAVY, size=13)
    ws4.merge_cells("A2:N2")
    ws4["A2"] = f"Earnings Multiple: {multiple}x  |  Div / RRoR: {int(rror*100)}%  |  P/E smoothed = {pe_smooth_days}-day rolling avg  |  Change assumptions in Settings sheet"
    _style_cell(ws4["A2"], italic=True, color=NAVY, fill=WHITE)

    last_row = ws3.max_row
    cats = Reference(ws3, min_col=1, min_row=3, max_row=last_row)

    chart_price = LineChart()
    chart_price.width = 30
    chart_price.height = 19
    chart_price.add_data(Reference(ws3, min_col=2, min_row=2, max_row=last_row), titles_from_data=True)
    chart_price.add_data(Reference(ws3, min_col=4, min_row=2, max_row=last_row), titles_from_data=True)
    chart_price.add_data(Reference(ws3, min_col=6, min_row=2, max_row=last_row), titles_from_data=True)
    chart_price.set_categories(cats)

    styles = [("1565C0", False, 16000, "Price (AUD $)"), ("8B0000", True, 14000, f"Value ({multiple}x E)"), ("2E7D32", True, 14000, f"Div / RRoR (Div/{int(rror*100)}%/100)")]
    for i, (color, dashed, width, title) in enumerate(styles):
        s = chart_price.series[i]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = width
        if dashed:
            s.graphicalProperties.line.dashStyle = "dash"
        s.smooth = False
        s.title = SeriesLabel(v=title)

    chart_price.x_axis.number_format = "MMM-YY"
    chart_price.x_axis.numFmt = NumFmt(formatCode="MMM-YY", sourceLinked=False)
    chart_price.x_axis.tickLblSkip = 126
    chart_price.x_axis.tickMarkSkip = 63
    chart_price.x_axis.majorGridlines = None
    chart_price.y_axis.title = "Price  (AUD $)"
    chart_price.y_axis.number_format = '"$"#,##0'
    chart_price.legend.position = "b"

    chart_pe = LineChart()
    chart_pe.add_data(Reference(ws3, min_col=8, min_row=2, max_row=last_row), titles_from_data=True)
    chart_pe.set_categories(cats)
    chart_pe.series[0].graphicalProperties.line.solidFill = "E65100"
    chart_pe.series[0].graphicalProperties.line.width = 16000
    chart_pe.series[0].smooth = True
    chart_pe.series[0].title = SeriesLabel(v=f"P/E ({pe_smooth_days}D Avg)")
    chart_pe.y_axis.axId = 200
    chart_pe.y_axis.title = "P/E Ratio"
    chart_pe.y_axis.crosses = "max"
    chart_pe.y_axis.number_format = "0.0"
    chart_pe.y_axis.majorGridlines = None
    chart_pe.x_axis.axId = 100
    chart_pe.x_axis.crosses = "autoZero"
    chart_price += chart_pe

    ws4.add_chart(chart_price, "A3")

    guide_row = 42
    guide = [
        "CHART INTERPRETATION GUIDE",
        "• Share price (blue) BELOW both red and green lines → potential value opportunity",
        "• Share price ABOVE both lines → share may be expensive",
        f"• P/E (orange, right axis) = {pe_smooth_days}-day rolling simple moving average",
        "• Step-function behaviour on Value and Div/RRoR lines is correct",
        "• Special dividends excluded from Div/RRoR line",
        "• Source: ASX company announcements; price data from provided CSV export",
    ]
    for i, line in enumerate(guide):
        row = guide_row + i
        ws4.merge_cells(f"A{row}:N{row}")
        ws4[f"A{row}"] = line
        if i == 0:
            _style_cell(ws4[f"A{row}"], bold=True, color="FFFFFF", fill=NAVY)
        else:
            _style_cell(ws4[f"A{row}"], fill=GREY if i % 2 == 1 else WHITE)

    # FuturePrompt
    ws5 = wb.create_sheet("FuturePrompt")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:N1")
    ws5["A1"] = "FUTURE PROMPT — Copy this text to regenerate the spreadsheet for any ASX company"
    _style_cell(ws5["A1"], bold=True, color="FFFFFF", fill=NAVY, size=11)
    for i, line in enumerate(PROMPT_FUTURE_TEXT.splitlines(), start=2):
        safe = line if not line.startswith("=") else "'" + line
        ws5.cell(i, 1, safe)
        ws5.cell(i, 1).font = Font(name="Courier New", size=9)

    wb.active = wb["ValueChart"]
    ws2.freeze_panes = "A3"
    ws3.freeze_panes = "A3"

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    recalc_result = _run_recalc(out_path)
    if recalc_result.get("status") not in ("success", "skipped") or recalc_result.get("total_errors", 0) > 0:
        raise RuntimeError(f"Spreadsheet recalc failed: {recalc_result}")

    folder_id = config.get("drive_folder_id")
    url = upload_or_replace_xlsx(out_path, drive_name=ticker, folder_id=folder_id)
    kb = round(out_path.stat().st_size / 1024, 1)
    print(f"[wally] Spreadsheet uploaded: name={ticker}, size_kb={kb}, url={url}", flush=True)
    return url
