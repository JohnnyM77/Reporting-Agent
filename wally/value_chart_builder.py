"""
wally/value_chart_builder.py
────────────────────────────
Universal value-chart spreadsheet builder for any market (ASX, US, Tokyo, etc.).
Reads a per-company YAML config from valuations/<ticker_slug>.yaml
and produces the canonical workbook (5 sheets, dual-axis chart, XML-patched
axis title boxes and legend).

Ticker → config mapping (universal, no .AX hard-coding):
  NHC.AX  → valuations/nhc_ax.yaml
  POOL    → valuations/pool.yaml
  FICO    → valuations/fico.yaml
  2914.T  → valuations/2914_t.yaml

If no config file exists, a starter config is auto-created and the workbook
is built with DATA REQUIRED placeholders where earnings history is absent.

Required workbook sheets (in order):
  1. Settings
  2. EarningsData
  3. PriceData
  4. ValueChart
  5. FuturePrompt

Called by:
  - scripts/build_value_chart.py  (CLI / agent invocation)
  - wally/main.py                 (auto-triggered on flagged companies)
  - wally/charts.py               (watchlist low-screen supplement)
"""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import yaml
from lxml import etree
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import NumFmt
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── XML namespaces ────────────────────────────────────────────────────────────
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
C    = f"{{{C_NS}}}"
A    = f"{{{A_NS}}}"

# ── Workbook palette (fixed) ──────────────────────────────────────────────────
NAVY       = "1F2D4E"
YELLOW     = "FFFF00"
BLUE_INPUT = "0000FF"
GREY       = "F2F2F2"
WHITE      = "FFFFFF"
BORDER_COL = "BBBBBB"


# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG LOADING
# ─────────────────────────────────────────────────────────────────────────────

# ── Exchange / currency inference ─────────────────────────────────────────────

_EXCHANGE_BY_SUFFIX: dict[str, str] = {
    "AX": "ASX",
    "T":  "TSE",
    "L":  "LSE",
    "HK": "HKEX",
    "SS": "SSE",
    "SZ": "SZSE",
}

_CURRENCY_BY_EXCHANGE: dict[str, str] = {
    "ASX":   "AUD",
    "TSE":   "JPY",
    "LSE":   "GBP",
    "HKEX":  "HKD",
    "SSE":   "CNY",
    "SZSE":  "CNY",
    "NYSE":  "USD",
    "NASDAQ": "USD",
}


def _ticker_slug(ticker: str) -> str:
    """Convert any ticker to a filesystem-safe slug.

    Examples::

        NHC.AX  → nhc_ax
        POOL    → pool
        FICO    → fico
        2914.T  → 2914_t
        MSFT    → msft
    """
    return ticker.lower().replace(".", "_")


def _infer_exchange(ticker: str) -> str:
    """Infer the primary listing exchange from a ticker string.

    Examples::

        NHC.AX  → ASX
        2914.T  → TSE
        POOL    → NASDAQ  (plain ticker — assumed US)
        MSFT    → NASDAQ
    """
    if "." in ticker:
        suffix = ticker.rsplit(".", 1)[1].upper()
        return _EXCHANGE_BY_SUFFIX.get(suffix, suffix)
    # Plain ticker (no suffix) → assume US NASDAQ/NYSE
    return "NASDAQ"


def _infer_currency(exchange: str) -> str:
    """Return the primary currency for a given exchange code."""
    return _CURRENCY_BY_EXCHANGE.get(exchange.upper(), "USD")


def _auto_create_starter_config(ticker: str, path: Path) -> dict[str, Any]:
    """Auto-create a minimal starter config when none exists.

    The created config is a valid starting point but will produce a workbook
    with "DATA REQUIRED" placeholders until earnings history is added.
    Logs the creation so the caller can see it happened.
    """
    exchange = _infer_exchange(ticker)
    currency = _infer_currency(exchange)
    cfg: dict[str, Any] = {
        "ticker":       ticker,
        "company_name": ticker.split(".")[0],
        "business":     "— update this field —",
        "exchange":     exchange,
        "currency":     currency,
        "stock_type":   "quality",
        "buy_multiple": 15,
        "sell_multiple": None,
        "rror":         0.04,
        "norm_eps":     None,
        "earnings":     [],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    header = (
        f"# {ticker} — Value Chart Config (auto-generated starter)\n"
        f"# Exchange: {exchange} | Currency: {currency}\n"
        f"# TODO: Update company_name, business, buy_multiple, rror, and earnings.\n"
        f"# Run: python scripts/build_value_chart.py {ticker}\n\n"
    )
    path.write_text(
        header + yaml.dump(cfg, default_flow_style=False, sort_keys=False),
        encoding="utf-8",
    )
    print(f"[wally] Auto-created starter config: {path}", flush=True)
    return cfg


def load_config(ticker_or_path: str) -> dict[str, Any]:
    """Load and validate a company config.

    Resolves in this order:
      1. Treat *ticker_or_path* as a literal file path.
      2. Auto-resolve to ``valuations/<ticker_slug>.yaml`` relative to repo root.
      3. If still not found, auto-create a minimal starter config (never raises
         ``FileNotFoundError`` — builds a DATA REQUIRED workbook instead).
    """
    p = Path(ticker_or_path)
    if not p.exists():
        # Try auto-resolve relative to repo root (2 levels up from this file)
        repo_root = Path(__file__).resolve().parents[1]
        p = repo_root / "valuations" / f"{_ticker_slug(ticker_or_path)}.yaml"
    if not p.exists():
        # Auto-create starter config rather than raising
        cfg = _auto_create_starter_config(ticker_or_path, p)
    else:
        cfg = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    _validate_config(cfg)
    return cfg


def _validate_config(cfg: dict) -> None:
    """Validate and normalise a company config dict.

    Hard requirements: ``ticker``, ``company_name``, ``buy_multiple``, ``rror``.
    Optional (defaults applied if absent):
      - ``earnings``     → empty list (workbook shows DATA REQUIRED placeholders)
      - ``sell_multiple``→ None (no sell line drawn)
      - ``exchange``     → inferred from ticker suffix
      - ``currency``     → inferred from exchange
    """
    required = ["ticker", "company_name", "buy_multiple", "rror"]
    missing  = [k for k in required if k not in cfg or cfg[k] is None]
    if missing:
        raise ValueError(f"Config missing required keys: {missing}")
    # Apply optional-field defaults
    if not cfg.get("sell_multiple"):
        cfg["sell_multiple"] = None
    if "earnings" not in cfg or cfg["earnings"] is None:
        cfg["earnings"] = []
    if not cfg.get("exchange"):
        cfg["exchange"] = _infer_exchange(cfg["ticker"])
    if not cfg.get("currency"):
        cfg["currency"] = _infer_currency(cfg["exchange"])


def _get(cfg: dict, *keys, default=None):
    """Safe nested dict getter: _get(cfg, 'chart', 'left_axis', 'max', default=8.0)"""
    v = cfg
    for k in keys:
        if not isinstance(v, dict):
            return default
        v = v.get(k, default)
        if v is None:
            return default
    return v


# ─────────────────────────────────────────────────────────────────────────────
#  XML PATCH HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _spPr_element(fill_hex: str, border_hex: str | None,
                  border_w: int) -> etree._Element:
    el = etree.Element(f"{C}spPr")
    sf = etree.SubElement(el, f"{A}solidFill")
    etree.SubElement(sf, f"{A}srgbClr").set("val", fill_hex)
    ln = etree.SubElement(el, f"{A}ln")
    if border_hex:
        ln.set("w", str(border_w))
        sf2 = etree.SubElement(ln, f"{A}solidFill")
        etree.SubElement(sf2, f"{A}srgbClr").set("val", border_hex)
    else:
        etree.SubElement(ln, f"{A}noFill")
    return el


def _rich_bold_tx(text: str, color: str = "000000",
                  size_pt: int = 9) -> etree._Element:
    tx    = etree.Element(f"{C}tx")
    rich  = etree.SubElement(tx, f"{C}rich")
    etree.SubElement(rich, f"{A}bodyPr")
    etree.SubElement(rich, f"{A}lstStyle")
    p     = etree.SubElement(rich, f"{A}p")
    pPr   = etree.SubElement(p, f"{A}pPr")
    defR  = etree.SubElement(pPr, f"{A}defRPr")
    defR.set("b",  "1")
    defR.set("sz", str(size_pt * 100))
    r     = etree.SubElement(p, f"{A}r")
    rPr   = etree.SubElement(r, f"{A}rPr")
    rPr.set("lang", "en-AU")
    rPr.set("b",    "1")
    rPr.set("sz",   str(size_pt * 100))
    sf    = etree.SubElement(rPr, f"{A}solidFill")
    etree.SubElement(sf, f"{A}srgbClr").set("val", color)
    etree.SubElement(r, f"{A}t").text = text
    return tx


def patch_chart_xml(xlsx_path: Path, cfg: dict) -> None:
    """
    Post-process chart XML inside the xlsx zip:
      1. Axis title boxes — fill / border / bold text (from chart.axis_title_box config)
      2. Secondary axis axPos forced to 'r'
      3. Legend box — fill / border (from chart.legend_box config)
    """
    atb    = _get(cfg, "chart", "axis_title_box") or {}
    lb     = _get(cfg, "chart", "legend_box")     or {}
    la_lbl = _get(cfg, "chart", "left_axis",  "label", default="Price (AUD $)")
    ra_lbl = _get(cfg, "chart", "right_axis", "label", default="P/E Ratio")

    fill_hex   = atb.get("fill",              "FFFF00")
    border_hex = atb.get("border_color",      "000000")
    border_w   = int(atb.get("border_width_emu", 9525))
    txt_color  = atb.get("text_color",        "000000")
    txt_size   = int(atb.get("font_size_pt",  9))

    leg_fill  = lb.get("fill",              "FFFFFF")
    leg_bdr   = lb.get("border_color",      "404040")
    leg_bdr_w = int(lb.get("border_width_emu", 12700))

    axis_labels = {"100": la_lbl, "200": ra_lbl}

    buf = io.BytesIO(xlsx_path.read_bytes())
    out = io.BytesIO()

    with zipfile.ZipFile(buf) as zin, \
         zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:

        for info in zin.infolist():
            data = zin.read(info.filename)

            if (info.filename.startswith("xl/charts/chart")
                    and info.filename.endswith(".xml")):

                root = etree.fromstring(data)

                # 1. Axis title boxes + axis position enforcement
                for ax in root.findall(f".//{C}valAx"):
                    ax_id_el = ax.find(f"{C}axId")
                    ax_id    = ax_id_el.get("val") if ax_id_el is not None else None

                    if ax_id == "200":
                        # Secondary value axis must be on the RIGHT
                        axpos = ax.find(f"{C}axPos")
                        if axpos is None:
                            axpos = etree.SubElement(ax, f"{C}axPos")
                        axpos.set("val", "r")
                    elif ax_id == "100":
                        # Primary value axis must be on the LEFT
                        axpos = ax.find(f"{C}axPos")
                        if axpos is None:
                            axpos = etree.SubElement(ax, f"{C}axPos")
                        axpos.set("val", "l")

                    title_el = ax.find(f"{C}title")
                    if title_el is None:
                        continue

                    lbl = axis_labels.get(ax_id or "", "")
                    if lbl:
                        old = title_el.find(f"{C}tx")
                        if old is not None:
                            idx = list(title_el).index(old)
                            title_el.remove(old)
                            title_el.insert(idx,
                                _rich_bold_tx(lbl, txt_color, txt_size))

                    for old in title_el.findall(f"{C}spPr"):
                        title_el.remove(old)
                    title_el.append(
                        _spPr_element(fill_hex, border_hex, border_w))

                # 2. Legend box
                legend_el = root.find(f".//{C}legend")
                if legend_el is not None:
                    for old in legend_el.findall(f"{C}spPr"):
                        legend_el.remove(old)
                    legend_el.append(
                        _spPr_element(leg_fill, leg_bdr, leg_bdr_w))

                data = etree.tostring(
                    root,
                    xml_declaration=True,
                    encoding="UTF-8",
                    standalone=True,
                )

            zout.writestr(info, data)

    xlsx_path.write_bytes(out.getvalue())


# ─────────────────────────────────────────────────────────────────────────────
#  SPREADSHEET HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


def _sc(cell, *, bold=False, color="000000", fill=None,
        align="left", wrap=False, size=9, italic=False):
    cell.font = Font(name="Arial", size=size, bold=bold,
                     color=color, italic=italic)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    cell.border = _thin_border()


def _band(ws, row: int, start_col: int, end_col: int):
    fill = GREY if row % 2 == 0 else WHITE
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill)


def _safe_half(ttm_vals: list[float]) -> list[float]:
    out, prev = [], None
    for v in ttm_vals:
        out.append(v / 2.0 if prev is None else max(0.0, v - prev / 2.0))
        prev = v
    return out


def _lookup_last(ts: pd.Timestamp,
                 dates: list[pd.Timestamp],
                 vals: list[float]) -> float | None:
    result = None
    for d, v in zip(dates, vals):
        if d <= ts:
            result = v
        else:
            break
    return result


# ─────────────────────────────────────────────────────────────────────────────
#  LIVE FUNDAMENTALS FETCH
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_live_fundamentals(
    ticker: str, cfg: dict
) -> tuple[list[dict], str, str, str]:
    """Fetch live earnings and dividend fundamentals for *ticker*.

    Uses Alpha Vantage for quarterly EPS history (requires ALPHAVANTAGE_API_KEY)
    and yfinance for dividend history (no key required).

    Returns:
        (earnings_history, earnings_source, div_source, warning_msg)

        *earnings_history* is a list of workbook-format dicts (may be empty).
        *earnings_source* / *div_source* are descriptive source labels.
        *warning_msg* is empty on success; set to a human-readable explanation
        when fundamentals could not be fetched (PRICE_ONLY fallback).
    """
    from .alphavantage import (
        build_workbook_earnings_history,
        fetch_earnings,
        get_api_key,
    )

    av_key = get_api_key()
    earnings_source = "none"
    div_source = "none"
    earnings_records = 0
    div_records = 0
    av_earnings_data = None
    div_series = None
    warnings: list[str] = []

    # ── Step 1: dividend history from yfinance (no API key needed) ────────────
    try:
        import yfinance as yf

        tk = yf.Ticker(ticker)
        divs = tk.dividends
        if divs is not None and not divs.empty:
            div_series = divs
            div_records = len(divs)
            div_source = "yfinance"
        else:
            div_source = "yfinance(none)"
    except Exception as e:
        print(f"[wally] {ticker} dividend fetch failed (yfinance): {e}", flush=True)
        warnings.append(f"dividend fetch failed: {e}")

    # ── Step 2: quarterly earnings from Alpha Vantage ─────────────────────────
    if av_key:
        try:
            av_earnings_data = fetch_earnings(ticker, av_key)
            earnings_records = len(av_earnings_data.quarterly)
            earnings_source = "alphavantage"
        except Exception as e:
            print(
                f"[wally] {ticker} earnings fetch failed (alphavantage): {e}",
                flush=True,
            )
            warnings.append(f"earnings fetch failed: {e}")
    else:
        print(
            f"[wally] {ticker} no ALPHAVANTAGE_API_KEY — skipping earnings fetch",
            flush=True,
        )
        warnings.append("no ALPHAVANTAGE_API_KEY set")

    # ── Step 3: build workbook earnings history ───────────────────────────────
    history: list[dict] = []
    if av_earnings_data is not None:
        history = build_workbook_earnings_history(av_earnings_data, div_series)

    # ── Step 4: log sources ───────────────────────────────────────────────────
    print(
        f"[wally] {ticker} price_source=yfinance"
        f" earnings_source={earnings_source}"
        f" dividends_source={div_source}",
        flush=True,
    )
    print(
        f"[wally] {ticker} earnings_records={earnings_records}"
        f" dividend_records={div_records}",
        flush=True,
    )

    if history:
        print(f"[wally] {ticker} build_status=FULL_BUILD", flush=True)
        return history, earnings_source, div_source, ""

    # ── Fallback: PRICE_ONLY ──────────────────────────────────────────────────
    reason = "; ".join(warnings) if warnings else "unknown reason"
    warning_msg = (
        "WARNING: Price data loaded, but live earnings/dividend data could not "
        f"be fetched. Reason: {reason}"
    )
    print(f"[wally] {ticker} build_status=PRICE_ONLY_BUILD", flush=True)
    return [], earnings_source, div_source, warning_msg


# ─────────────────────────────────────────────────────────────────────────────
#  PRICE DATA
# ─────────────────────────────────────────────────────────────────────────────

def get_price_data(cfg: dict,
                   price_csv_path: str | None = None) -> pd.DataFrame:
    """
    Return weekly (or daily) price DataFrame with columns [Date, Close].
    Priority: 1) explicit CSV path  2) yfinance  3) synthetic placeholder
    """
    ticker     = cfg["ticker"]
    freq       = _get(cfg, "price_data", "frequency", default="weekly")
    resample   = _get(cfg, "price_data", "resample_day", default="W-FRI")
    start_date = _get(cfg, "price_data", "start_date", default="2016-01-01")

    # 1 — CSV supplied by caller (e.g. from marketindex export)
    if price_csv_path and Path(price_csv_path).exists():
        df = pd.read_csv(price_csv_path)
        # Normalise: accept YYYYMMDD integers or parseable date strings
        df["Date"]  = pd.to_datetime(df["Date"].astype(str),
                                      format="%Y%m%d", errors="coerce")
        df["Close"] = pd.to_numeric(df["Close"], errors="coerce")
        df = (df[["Date", "Close"]]
              .dropna()
              .sort_values("Date")
              .query("Date >= @start_date"))
        if freq == "weekly":
            df = (df.set_index("Date")
                    .resample(resample).last()
                    .dropna()
                    .reset_index())
        print(f"[vcb] CSV loaded: {len(df)} rows ({freq})")
        return df

    # 2 — yfinance
    try:
        import yfinance as yf
        tk   = yf.Ticker(ticker)
        hist = tk.history(start=start_date, interval="1d",
                          auto_adjust=True)
        if hist.empty or "Close" not in hist:
            raise ValueError("Empty history")
        df = hist[["Close"]].reset_index()
        df.columns = ["Date", "Close"]
        df["Date"] = pd.to_datetime(df["Date"]).dt.tz_localize(None)
        df = df.dropna().sort_values("Date")
        if freq == "weekly":
            df = (df.set_index("Date")
                    .resample(resample).last()
                    .dropna()
                    .reset_index())
        print(f"[vcb] yfinance: {len(df)} rows ({freq})")
        return df
    except Exception as e:
        print(f"[vcb] yfinance failed ({e}) — synthetic placeholder data")

    # 3 — Synthetic placeholder (shape approximates a typical ASX stock)
    dates  = pd.date_range(start_date, periods=530, freq=resample)
    n      = len(dates)
    idx    = np.arange(n) / n
    prices = np.maximum(0.5, 2.0 + 3.0 * np.sin(idx * np.pi)
                         + np.random.default_rng(42).normal(0, 0.1, n))
    return pd.DataFrame({"Date":  dates,
                          "Close": np.round(prices, 2)})


# ─────────────────────────────────────────────────────────────────────────────
#  SHEET BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def _build_settings(wb: Workbook, cfg: dict) -> None:
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 62

    ticker   = cfg["ticker"]
    name     = cfg["company_name"]
    biz      = cfg.get("business", "")
    bmult    = cfg["buy_multiple"]
    smult    = cfg.get("sell_multiple")
    rror     = float(cfg["rror"])
    norm     = cfg.get("norm_eps")
    stype    = cfg.get("stock_type", "quality")
    exchange = cfg.get("exchange", "")
    currency = cfg.get("currency", "USD")

    # Exchange label for titles: "ASX: NHC" or "NASDAQ: POOL" or "TSE: 2914"
    ticker_display = ticker.split(".")[0]
    exchange_label = f"{exchange}: " if exchange else ""

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = f"{name} ({exchange_label}{ticker_display}) — Value Analysis — Settings"
    _sc(ws["A1"], bold=True, color="FFFFFF", fill=NAVY, align="center", size=13)
    ws.row_dimensions[1].height = 28

    # Section headers
    for r, txt in {
        3:  "COMPANY DETAILS",
        7:  "VALUATION ASSUMPTIONS",
        14: "ANALYSIS PERIOD",
        17: "HOW TO REFRESH SHARE PRICE DATA",
    }.items():
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"] = txt
        _sc(ws[f"A{r}"], bold=True, color="FFFFFF", fill=NAVY)
        ws.row_dimensions[r].height = 20

    for r, lbl, val in [
        (4, "Company Name",                              name),
        (5, f"{exchange} Ticker" if exchange else "Ticker", ticker),
        (6, "Business",                                  biz),
    ]:
        ws[f"A{r}"] = lbl;  ws[f"B{r}"] = val
        for c in ["A", "B", "C"]: _sc(ws[f"{c}{r}"])

    # Valuation inputs
    rows_8_9_10 = [
        (8,  "Buy Multiple (P/E)",             bmult,  "0",
             f"SLOW TO BUY: price < TTM EPS × {bmult} ÷ 100"),
        (9,  "Sell Multiple (P/E)",             smult or "",  "0",
             (f"EVEN SLOWER TO SELL: price > TTM EPS × {smult} ÷ 100"
              if smult else "Not set — no sell line on chart")),
        (10, "Required Rate of Return (RRoR)", rror,  "0.00%",
             f"Min dividend yield you require ({int(rror*100)}%) — drives Div/RRoR line"),
    ]
    if norm is not None:
        rows_8_9_10.append(
            (11, "Normalised Through-Cycle EPS (¢)", norm, "0",
             f"Mid-cycle EPS estimate ({stype} stock). "
             "Adjust as your earnings view evolves."))
    for r, lbl, val, fmt, note in rows_8_9_10:
        ws[f"A{r}"] = lbl;  ws[f"B{r}"] = val;  ws[f"C{r}"] = note
        _sc(ws[f"A{r}"])
        _sc(ws[f"B{r}"], color=BLUE_INPUT, fill=YELLOW)
        _sc(ws[f"C{r}"], italic=True, color="555555")
        ws[f"B{r}"].number_format = fmt

    # Normalised price reference formulas (if applicable)
    if norm is not None:
        for r, lbl, formula, note in [
            (12, "→ Normalised Buy Price ($)",  "=B11*B8/100",
             f"= Norm EPS × Buy Multiple ÷ 100  →  ${norm*bmult/100:.2f}"),
            (13, "→ Normalised Sell Price ($)", "=B11*B9/100",
             f"= Norm EPS × Sell Multiple ÷ 100  →  ${norm*(smult or 0)/100:.2f}"),
        ]:
            ws[f"A{r}"] = lbl;  ws[f"B{r}"] = formula;  ws[f"C{r}"] = note
            for c in ["A", "B", "C"]:
                _sc(ws[f"{c}{r}"], italic=True, color="333333")
            ws[f"B{r}"].number_format = '"$"#,##0.00'

    ws["A15"], ws["B15"] = "From (approx.)", "Jan 2016"
    ws["A16"], ws["B16"] = "To (approx.)",   "Latest available"
    for r in [15, 16]:
        for c in ["A", "B"]: _sc(ws[f"{c}{r}"])

    src     = _get(cfg, "price_data", "source_hint",
                   default="marketindex.com.au — export daily CSV (Date as YYYYMMDD)")
    formula = _get(cfg, "price_data", "stockhistory_formula", default="")
    for i, line in enumerate([
        f"1. Go to {src}",
        "2. Ensure CSV columns: Date, Open, High, Low, Close, Volume",
        "3. In PriceData sheet: paste Date → col A, Close → col B from row 3 onwards",
        "4. PriceData uses WEEKLY Friday close (resampled). Paste daily — agent resamples.",
        "5. Chart updates automatically.",
        f"6. Microsoft 365: {formula}",
    ], start=18):
        ws.merge_cells(f"A{i}:C{i}")
        ws[f"A{i}"] = line
        _sc(ws[f"A{i}"], italic=True, color="666666")

    # Cyclical warning block
    if stype == "cyclical" and norm:
        ws.merge_cells("A24:C24")
        ws["A24"] = "⚠  CYCLICAL STOCK WARNING — READ BEFORE CHANGING MULTIPLES"
        _sc(ws["A24"], bold=True, color="FFFFFF", fill=NAVY)
        ws.row_dimensions[24].height = 20
        warn_lines = [
            (f"{name} is CYCLICAL — 'low P/E = cheap' is INVERTED at peak earnings.", True),
            ("At peak earnings the sell line reaches an unreachable number. Use the Normalised Sell price as your real ceiling.", False),
            ("At trough earnings the buy line drops too low. Use the Normalised Buy price as your floor.", False),
            (f"RULE: Slow to buy ({bmult}×). Even SLOWER to sell ({smult}×). "
             "Only act when the normalised price is clearly in zone.", True),
            (f"Normalised Buy  = ${norm*bmult/100:.2f}  |  "
             f"Normalised Sell = ${norm*(smult or 0)/100:.2f}  (update Settings!B11 as view changes)", False),
        ]
        for i, (line, is_rule) in enumerate(warn_lines, start=25):
            ws.merge_cells(f"A{i}:C{i}")
            ws[f"A{i}"] = line
            _sc(ws[f"A{i}"],
                bold=is_rule,
                color="8B2500" if is_rule else "333333",
                fill="FFF3E0" if i % 2 == 0 else "FFFFFF",
                wrap=True)
            ws.row_dimensions[i].height = 22


def _build_earnings(wb: Workbook, cfg: dict) -> None:
    ws2 = wb.create_sheet("EarningsData")
    ws2.sheet_view.showGridLines = False
    ticker   = cfg["ticker"].split(".")[0]
    name     = cfg["company_name"]
    bmult    = cfg["buy_multiple"]
    smult    = cfg.get("sell_multiple")
    exchange = cfg.get("exchange", "")
    currency = cfg.get("currency", "USD")

    exchange_label = f"{exchange}: " if exchange else ""

    headers = [
        "Ann. Date", "Period",
        f"Half EPS\n({currency} ¢)", f"TTM EPS\n({currency} ¢)",
        f"Half Div\n({currency} ¢)", f"TTM Div\n({currency} ¢)",
        f"Value Buy\n({bmult}× E)",
    ]
    if smult:
        headers.append(f"Value Sell\n({smult}× E)")
    headers += ["Div / RRoR\n(÷RRoR÷100)", "Notes"]

    widths = ([14, 30, 11, 11, 10, 10, 14]
              + ([15] if smult else [])
              + [14, 52])
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(64+i)].width = w

    ws2.merge_cells(f"A1:{chr(64+len(headers))}1")
    ws2["A1"] = f"{name} ({exchange_label}{ticker}) — Earnings & Dividend Data"
    _sc(ws2["A1"], bold=True, color="FFFFFF", fill=NAVY, size=13)
    ws2.row_dimensions[1].height = 28

    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=c, value=h)
        _sc(cell, bold=True, color="FFFFFF", fill=NAVY, align="center", wrap=True)
    ws2.row_dimensions[2].height = 42

    earnings = cfg.get("earnings") or []

    # If no earnings data, show a visible DATA REQUIRED / PRICE_ONLY warning and exit early
    if not earnings:
        from openpyxl.utils import get_column_letter
        ncols = len(headers)
        end_col = get_column_letter(ncols)
        price_only_msg = cfg.get("_price_only_warning", "")
        if price_only_msg:
            ws2.merge_cells(f"A3:{end_col}3")
            ws2["A3"] = price_only_msg
            _sc(ws2["A3"], bold=True, color="8B2500", fill="FFF3E0", wrap=True)
            ws2.row_dimensions[3].height = 40
            ws2.merge_cells(f"A4:{end_col}5")
            ws2["A4"] = (
                "Add earnings history to "
                f"valuations/{_ticker_slug(cfg['ticker'])}.yaml "
                "or set ALPHAVANTAGE_API_KEY to enable valuation lines."
            )
            _sc(ws2["A4"], bold=False, color="8B0000", fill="FFF3E0", wrap=True)
            ws2.row_dimensions[4].height = 30
        else:
            ws2.merge_cells(f"A3:{end_col}4")
            ws2["A3"] = (
                "⚠ DATA REQUIRED — Add earnings history to "
                f"valuations/{_ticker_slug(cfg['ticker'])}.yaml to enable valuation lines"
            )
            _sc(ws2["A3"], bold=True, color="8B0000", fill="FFF3E0", wrap=True)
            ws2.row_dimensions[3].height = 40
        return
    ttm_eps  = [float(e["ttm_eps"]) for e in earnings]
    ttm_div  = [float(e["ttm_div"]) for e in earnings]
    half_eps = _safe_half(ttm_eps)
    half_div = _safe_half(ttm_div)

    for idx, e in enumerate(earnings, start=3):
        ann_date = (e["date"] if isinstance(e["date"], date)
                    else datetime.strptime(str(e["date"]), "%Y-%m-%d").date())
        r = idx
        col = 1
        ws2.cell(r, col, ann_date);              col += 1
        ws2.cell(r, col, e["period"]);           col += 1
        ws2.cell(r, col, round(half_eps[idx-3], 1)); col += 1
        ws2.cell(r, col, float(e["ttm_eps"]));   col += 1
        ws2.cell(r, col, round(half_div[idx-3], 1)); col += 1
        ws2.cell(r, col, float(e["ttm_div"]));   col += 1
        ws2.cell(r, col, f"=D{r}*Settings!$B$8/100");  col += 1
        if smult:
            ws2.cell(r, col, f"=D{r}*Settings!$B$9/100"); col += 1
        ws2.cell(r, col, f"=F{r}/Settings!$B$10/100"); col += 1
        ws2.cell(r, col, str(e.get("notes", "")))

        for c in range(1, len(headers) + 1):
            _sc(ws2.cell(r, c),
                align="center" if c not in (1, 2, len(headers)) else "left")
        _band(ws2, r, 1, len(headers))

    for r in range(3, ws2.max_row + 1):
        ws2.cell(r, 1).number_format = "DD-MMM-YYYY"
        for c in [3, 4, 5, 6]:
            ws2.cell(r, c).number_format = "#,##0.0"
        buy_col  = 7
        sell_col = 8 if smult else None
        drror_col = 9 if smult else 8
        ws2.cell(r, buy_col).number_format = '"$"#,##0.00'
        if sell_col:
            ws2.cell(r, sell_col).number_format = '"$"#,##0.00'
        ws2.cell(r, drror_col).number_format = '"$"#,##0.00'

    leg = ws2.max_row + 2
    ws2.merge_cells(f"A{leg}:{chr(64+len(headers))}{leg}")
    ws2[f"A{leg}"] = (
        f"TTM EPS = trailing 12-month underlying EPS | "
        f"Value Buy = TTM EPS × {bmult}× ÷ 100"
        + (f" | Value Sell = TTM EPS × {smult}× ÷ 100" if smult else "")
        + " | Div/RRoR = TTM Div ÷ RRoR ÷ 100 | "
        "Ordinary dividends only — specials excluded"
    )
    _sc(ws2[f"A{leg}"], italic=True, color="666666", wrap=True)
    ws2.row_dimensions[leg].height = 30


def _build_price_data(wb: Workbook, cfg: dict,
                      price_df: pd.DataFrame) -> None:
    ws3 = wb.create_sheet("PriceData")
    ws3.sheet_view.showGridLines = False
    ticker     = cfg["ticker"].split(".")[0]
    name       = cfg["company_name"]
    bmult      = cfg["buy_multiple"]
    smult      = cfg.get("sell_multiple")
    rror       = float(cfg["rror"])
    pe_days    = int(_get(cfg, "chart", "pe_smooth_days", default=60))
    resample_r = _get(cfg, "price_data", "resample_day", default="W-FRI")
    pe_window  = 12 if "W" in resample_r else pe_days
    exchange   = cfg.get("exchange", "")
    currency   = cfg.get("currency", "USD")

    exchange_label = f"{exchange}: " if exchange else ""

    cols = [
        ("Date\n(Fri)", 13), (f"Price\n({currency} $)", 11),
        ("TTM EPS\n(¢)", 10), (f"Value Buy\n({bmult}× E)", 14),
    ]
    if smult:
        cols.append((f"Value Sell\n({smult}× E)", 15))
    cols += [
        ("TTM Div\n(¢)", 10), ("Div / RRoR\n($)", 13),
        ("P/E\nRatio", 9), (f"P/E {pe_days}D\nAvg", 13),
    ]
    for i, (h, w) in enumerate(cols, start=1):
        ws3.column_dimensions[chr(64+i)].width = w

    ws3.merge_cells(f"A1:{chr(64+len(cols))}1")
    ws3["A1"] = f"{name} ({exchange_label}{ticker}) — Weekly Price & Value Data"
    _sc(ws3["A1"], bold=True, color="FFFFFF", fill=NAVY, size=11)
    for c, (h, _) in enumerate(cols, start=1):
        cell = ws3.cell(2, c, h)
        _sc(cell, bold=True, color="FFFFFF", fill=NAVY, align="center", wrap=True)
    ws3.row_dimensions[2].height = 42

    _earnings = cfg.get("earnings") or []
    earn_dates = [pd.Timestamp(
        e["date"] if isinstance(e["date"], date)
        else datetime.strptime(str(e["date"]), "%Y-%m-%d").date()
    ) for e in _earnings]
    earn_eps = [float(e["ttm_eps"]) for e in _earnings]
    earn_div = [float(e["ttm_div"]) for e in _earnings]

    rows_data = []
    for _, row in price_df.iterrows():
        d     = pd.Timestamp(row["Date"])
        close = float(row["Close"])
        eps   = _lookup_last(d, earn_dates, earn_eps)
        div   = _lookup_last(d, earn_dates, earn_div)
        v_buy  = (eps * bmult  / 100) if eps is not None else None
        v_sell = (eps * smult  / 100) if (smult and eps is not None) else None
        drror  = (div / rror / 100)   if (div is not None and rror > 0) else None
        pe     = (close / (eps / 100)) if (eps is not None and eps > 0) else np.nan
        rows_data.append((d.to_pydatetime().date(),
                          close, eps, v_buy, v_sell, div, drror, pe))

    pe_s      = pd.Series([r[7] for r in rows_data], dtype=float)
    pe_smooth = pe_s.rolling(window=pe_window,
                              min_periods=max(4, pe_window//3)).mean().round(2)

    for idx, r in enumerate(rows_data, start=3):
        col_vals = [r[0], r[1], r[2], r[3]]
        if smult:
            col_vals.append(r[4])
        col_vals += [
            r[5], r[6],
            None if pd.isna(r[7]) else float(r[7]),
            None if pd.isna(pe_smooth.iloc[idx-3]) else float(pe_smooth.iloc[idx-3]),
        ]
        for c, v in enumerate(col_vals, start=1):
            ws3.cell(idx, c, v)
            _sc(ws3.cell(idx, c), align="center")
        _band(ws3, idx, 1, len(cols))

    for r in range(3, ws3.max_row + 1):
        ws3.cell(r, 1).number_format = "DD-MMM-YYYY"
        ws3.cell(r, 2).number_format = '"$"#,##0.00'
        ws3.cell(r, 3).number_format = "#,##0.0"
        ws3.cell(r, 4).number_format = '"$"#,##0.00'
        c = 5
        if smult:
            ws3.cell(r, c).number_format = '"$"#,##0.00'
            c += 1
        ws3.cell(r, c).number_format = "#,##0.0"    # div cents
        ws3.cell(r, c+1).number_format = '"$"#,##0.00'  # div/rror
        ws3.cell(r, c+2).number_format = "0.0"       # pe
        ws3.cell(r, c+3).number_format = "0.0"       # pe smooth


def _build_chart(wb: Workbook, cfg: dict, ws3) -> None:
    ws4 = wb.create_sheet("ValueChart")
    ws4.sheet_view.showGridLines = False
    ticker   = cfg["ticker"].split(".")[0]
    name     = cfg["company_name"]
    bmult    = cfg["buy_multiple"]
    smult    = cfg.get("sell_multiple")
    rror     = float(cfg["rror"])
    pe_d     = int(_get(cfg, "chart", "pe_smooth_days", default=60))
    exchange = cfg.get("exchange", "")

    exchange_label = f"{exchange}: " if exchange else ""

    # Chart subtitle row
    ws4.merge_cells("A1:N1")
    ws4["A1"] = (
        f"{name} ({exchange_label}{ticker})  —  Buy / Sell Zone Chart  |  "
        f"Buy {bmult}×  (Slow to buy)"
        + (f"  |  Sell {smult}×  (Even slower to sell)" if smult else "")
    )
    _sc(ws4["A1"], bold=True, color="FFFFFF", fill=NAVY, align="center", size=12)
    ws4.row_dimensions[1].height = 26

    norm = cfg.get("norm_eps")
    ws4.merge_cells("A2:N2")
    ws4["A2"] = (
        (f"Normalised EPS: {norm}¢  →  Buy ${norm*bmult/100:.2f}  /  Sell ${norm*(smult or 0)/100:.2f}  |  "
         if norm else "")
        + f"RRoR {int(rror*100)}%  |  P/E {pe_d}-day avg  |  Change in Settings sheet"
    )
    _sc(ws4["A2"], italic=True, color=NAVY, fill=WHITE)

    last_row = ws3.max_row
    cats     = Reference(ws3, min_col=1, min_row=3, max_row=last_row)

    # ── Primary chart (Price → left axis) ────────────────────────────────────
    ch = LineChart()
    # Assign explicit axis IDs to avoid collision with secondary chart.
    # openpyxl LineChart defaults: y_axis.axId=200, x_axis.axId=100.
    # We override them so the primary value axis uses ID 100 (left) and the
    # shared category axis uses ID 10, leaving ID 200 exclusively for the
    # secondary value axis (right).
    ch.y_axis.axId   = 100   # primary value axis — LEFT
    ch.x_axis.axId   = 10    # category axis — shared with secondary chart
    ch.width  = int(_get(cfg, "chart", "dimensions", "width",  default=34))
    ch.height = int(_get(cfg, "chart", "dimensions", "height", default=20))
    ch.add_data(Reference(ws3, min_col=2, min_row=2, max_row=last_row),
                titles_from_data=True)
    ch.set_categories(cats)

    lines_cfg = _get(cfg, "chart", "lines") or {}

    def _apply_line(series, key: str):
        lc      = lines_cfg.get(key, {})
        color   = lc.get("color",      "1565C0")
        width   = int(lc.get("width_emu", 9525))
        dashed  = lc.get("style", "solid") == "dash"
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width     = width
        if dashed:
            series.graphicalProperties.line.dashStyle = "dash"
        series.smooth        = False
        series.marker.symbol = "none"

    _apply_line(ch.series[0], "price")
    ch.series[0].title = SeriesLabel(v="Price (AUD $)")

    # Left axis
    la = _get(cfg, "chart", "left_axis") or {}
    ch.y_axis.title         = la.get("label",      "Price (AUD $)")
    ch.y_axis.number_format = la.get("num_fmt",    '"$"#,##0.00')
    ch.y_axis.majorUnit     = float(la.get("major_unit", 1.0))
    ch.y_axis.scaling.min   = float(la.get("min",  0.0))
    ch.y_axis.scaling.max   = float(la.get("max",  10.0))
    ch.y_axis.majorGridlines = None
    ch.y_axis.crossAx        = 10    # primary valAx crosses catAx (id=10)
    ch.x_axis.crossAx        = 100   # catAx crosses primary valAx (id=100)

    xa = _get(cfg, "chart", "x_axis") or {}
    ch.x_axis.number_format  = xa.get("num_fmt",          "MMM-YY")
    ch.x_axis.numFmt          = NumFmt(
        formatCode=xa.get("num_fmt", "MMM-YY"), sourceLinked=False)
    ch.x_axis.tickLblSkip    = int(xa.get("tick_label_skip", 26))
    ch.x_axis.tickMarkSkip   = int(xa.get("tick_mark_skip",  26))
    ch.x_axis.majorGridlines = None

    lp  = _get(cfg, "chart", "legend_position", default="t")
    lov = _get(cfg, "chart", "legend_overlay",  default=True)
    ch.legend.position = lp
    ch.legend.overlay  = lov

    # ── Secondary chart (value lines + P/E → right axis) ─────────────────────
    sec = LineChart()

    # Determine column numbers for value series
    has_sell = bool(smult)
    # Cols: 2=Price, 3=EPS, 4=ValueBuy, 5=ValueSell(if smult), then Div, DivRRoR, PE, PE_smooth
    value_buy_col  = 4
    value_sell_col = 5 if has_sell else None
    div_rror_col   = 7 if has_sell else 6
    pe_col         = 9 if has_sell else 8

    sec_cols = [value_buy_col]
    if has_sell:
        sec_cols.append(value_sell_col)
    sec_cols += [div_rror_col, pe_col]

    for col in sec_cols:
        sec.add_data(Reference(ws3, min_col=col, min_row=2, max_row=last_row),
                     titles_from_data=True)
    sec.set_categories(cats)

    sec_series_keys = ["value_buy"]
    if has_sell:
        sec_series_keys.append("value_sell")
    sec_series_keys += ["div_rror", "pe_smooth"]

    sec_titles = [f"Value Buy ({bmult}× TTM EPS)"]
    if has_sell:
        sec_titles.append(f"Value Sell ({smult}× TTM EPS)")
    sec_titles += [f"Div / RRoR ({int(rror*100)}%)", f"P/E ({pe_d}D Avg)"]

    for i, (key, title) in enumerate(zip(sec_series_keys, sec_titles)):
        _apply_line(sec.series[i], key)
        sec.series[i].title = SeriesLabel(v=title)

    ra = _get(cfg, "chart", "right_axis") or {}
    sec.y_axis.axId          = 200   # secondary value axis — RIGHT (unique, != primary's 100)
    sec.y_axis.title         = ra.get("label",      "P/E Ratio")
    sec.y_axis.crosses       = "max"
    sec.y_axis.crossAx       = 10    # secondary valAx crosses shared catAx (id=10)
    sec.y_axis.number_format = ra.get("num_fmt",    "0.0")
    sec.y_axis.majorUnit     = float(ra.get("major_unit", 2.0))
    sec.y_axis.scaling.min   = float(ra.get("min",  0.0))
    sec.y_axis.scaling.max   = float(ra.get("max",  20.0))
    sec.y_axis.majorGridlines = None
    sec.x_axis.axId           = 10    # shares the primary chart's catAx (id=10)
    sec.x_axis.crosses        = "autoZero"
    sec.x_axis.crossAx        = 100   # catAx crosses primary valAx (id=100)

    ch += sec
    ws4.add_chart(ch, "A3")

    # Interpretation guide
    guide_row = 45
    lines_guide = [
        ("CHART INTERPRETATION", "header"),
        (f"SLOW TO BUY: Price (blue) < Value Buy (dark green dashed, {bmult}× TTM EPS) → consider buying.", "buy"),
    ]
    if smult:
        lines_guide.append(
            (f"EVEN SLOWER TO SELL: Price > Value Sell (red dashed, {smult}× TTM EPS) → only then consider trimming.", "sell"))
    if norm:
        lines_guide += [
            (f"⚠  CYCLICAL: At peak earnings Value Sell line exceeds any realistic market price. "
             f"Use Normalised Sell (${norm*(smult or 0)/100:.2f}) as real ceiling.", "warn"),
            (f"⚠  AT TROUGH: Value Buy drops too low. "
             f"Use Normalised Buy (${norm*bmult/100:.2f}) as floor.", "warn"),
        ]
    lines_guide += [
        (f"P/E (olive, right axis, max {ra.get('max', 20)}×): context for whether earnings multiple is stretched.", "normal"),
        ("Div/RRoR (light green): price ≤ this line → dividend yield covers your return hurdle.", "normal"),
    ]
    for i, (line, style) in enumerate(lines_guide):
        row = guide_row + i
        ws4.merge_cells(f"A{row}:N{row}")
        ws4[f"A{row}"] = line
        if style == "header":
            _sc(ws4[f"A{row}"], bold=True, color="FFFFFF", fill=NAVY)
        elif style == "buy":
            _sc(ws4[f"A{row}"], bold=True, color="1A5E20", fill=GREY)
        elif style == "sell":
            _sc(ws4[f"A{row}"], bold=True, color="C0392B", fill=GREY)
        elif style == "warn":
            _sc(ws4[f"A{row}"], bold=True, color="8B0000", fill="FFF3E0")
        else:
            _sc(ws4[f"A{row}"], fill=WHITE if i % 2 == 0 else GREY, color="333333")
        ws4.row_dimensions[row].height = 20


def _build_future_prompt(wb: Workbook, cfg: dict) -> None:
    ws5 = wb.create_sheet("FuturePrompt")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 110
    ws5.merge_cells("A1:B1")
    ws5["A1"] = ("FUTURE PROMPT — Copy this text to regenerate or "
                 "adapt the spreadsheet for any company on any market")
    _sc(ws5["A1"], bold=True, color="FFFFFF", fill=NAVY, size=11)

    ticker  = cfg["ticker"]
    name    = cfg["company_name"]
    bmult   = cfg["buy_multiple"]
    smult   = cfg.get("sell_multiple", "None")
    rror    = float(cfg["rror"])
    norm    = cfg.get("norm_eps", "N/A")
    stype   = cfg.get("stock_type", "quality")
    la      = _get(cfg, "chart", "left_axis")  or {}
    ra      = _get(cfg, "chart", "right_axis") or {}
    xa      = _get(cfg, "chart", "x_axis")     or {}
    lns     = _get(cfg, "chart", "lines")      or {}
    atb     = _get(cfg, "chart", "axis_title_box") or {}
    lb      = _get(cfg, "chart", "legend_box")     or {}

    def lw(key): return lns.get(key, {}).get("width_emu", "—")
    def lc(key): return lns.get(key, {}).get("color", "—")
    def ls(key): return lns.get(key, {}).get("style", "—")

    exchange = cfg.get("exchange", "")
    currency = cfg.get("currency", "USD")
    slug     = _ticker_slug(ticker)
    prompt = f"""=== VALUE CHART — AGENT PROMPT v3 (supports ASX, US, TSE and all markets) ===
To regenerate: run  python scripts/build_value_chart.py {ticker}
To create new company: copy valuations/{ticker.lower().replace('.','_')}.yaml,
  update all fields, then run  python scripts/build_value_chart.py NEW_TICKER

Reference template: outputs/NHC_ASX_Value_Analysis_v3.xlsx  (cyclical example)
Builder module:     wally/value_chart_builder.py
Config schema:      valuations/nhc_ax.yaml  (full annotated example)

━━━━ COMPANY ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  ticker:       {ticker}    (slug → valuations/{slug}.yaml)
  company_name: {name}
  exchange:     {exchange}  (ASX | NASDAQ | NYSE | TSE | LSE | …)
  currency:     {currency}  (AUD | USD | JPY | GBP | …)
  business:     {cfg.get("business", "")}
  stock_type:   {stype}  ("cyclical" | "quality" | "growth" | "income")

━━━━ VALUATION INPUTS (Settings sheet yellow cells) ━━━━━━━━━━━━━━
  B8  buy_multiple:  {bmult}   — Slow to buy
  B9  sell_multiple: {smult}   — Even slower to sell (null = no sell line)
  B10 rror:          {rror}    — Required rate of return ({int(rror*100)}%)
  B11 norm_eps:      {norm}    — Normalised through-cycle EPS cents (null for quality stocks)
  B12 → Normalised Buy  Price: =B11*B8/100   (formula, display only)
  B13 → Normalised Sell Price: =B11*B9/100   (formula, display only)

━━━━ WORKBOOK STRUCTURE (5 sheets) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  1. Settings      — company + yellow input cells + cyclical warnings (if stock_type=cyclical)
  2. EarningsData  — half-year periods; Value Buy =D×B8/100, Value Sell =D×B9/100, Div/RRoR =F/B10/100
  3. PriceData     — WEEKLY prices (Friday close, pandas resample("W-FRI"))
                     ~530 rows over 10 years; far cleaner chart line than daily (~2600 rows)
                     Columns: Date(Fri)|Price|TTM_EPS|ValueBuy|ValueSell*|TTM_Div|Div/RRoR|P/E|P/E_smooth
                     (*ValueSell column omitted if sell_multiple is null)
                     P/E smooth: rolling(window=12, min_periods=4).mean()  [12 wks ≈ {int(_get(cfg,"chart","pe_smooth_days",default=60))} days]
  4. ValueChart    — dual-axis chart (see CHART SPEC below)
  5. FuturePrompt  — this text, Courier New 9pt, column A width=110

━━━━ CHART SPECIFICATION ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Positioned at: A3  |  width={_get(cfg,"chart","dimensions","width",default=34)}  height={_get(cfg,"chart","dimensions","height",default=20)} (openpyxl units)

  SERIES & AXIS ASSIGNMENT:
  ┌──────────────────────────────────┬────────────────┬────────────┬────────────────────────┐
  │ Series                           │ Axis           │ Colour hex │ Width (EMU) / Style    │
  ├──────────────────────────────────┼────────────────┼────────────┼────────────────────────┤
  │ Price (AUD $)                    │ PRIMARY LEFT   │ {lc("price")}   │ {lw("price")} / {ls("price")}              │
  │ Value Buy  ({bmult}× TTM EPS)           │ SECONDARY RIGHT│ {lc("value_buy")}   │ {lw("value_buy")} / {ls("value_buy")}            │
  │ Value Sell ({smult}× TTM EPS)          │ SECONDARY RIGHT│ {lc("value_sell")}   │ {lw("value_sell")} / {ls("value_sell")}            │
  │ Div / RRoR ({int(rror*100)}%)            │ SECONDARY RIGHT│ {lc("div_rror")}   │ {lw("div_rror")} / {ls("div_rror")}            │
  │ P/E ({int(_get(cfg,"chart","pe_smooth_days",default=60))}D Avg)                    │ SECONDARY RIGHT│ {lc("pe_smooth")}   │ {lw("pe_smooth")} / {ls("pe_smooth")}              │
  └──────────────────────────────────┴────────────────┴────────────┴────────────────────────┘
  EMU reference: 12700 = 1pt | 9525 = 0.75pt | 19050 = 1.5pt | 15875 = 1.25pt
  All series: marker.symbol = "none"

  PRIMARY AXIS (left):
    title: "{la.get("label","Price (AUD $)")}"   min:{la.get("min",0)}  max:{la.get("max",10)}  majorUnit:{la.get("major_unit",1)}
    numFmt: {la.get("num_fmt",'"$"#,##0.00')}    gridlines: None

  SECONDARY AXIS (right):
    axId: 200   crosses: "max"   axPos: "r"  ← MUST be set via XML patch (see below)
    title: "{ra.get("label","P/E Ratio")}"   min:{ra.get("min",0)}  max:{ra.get("max",20)}  majorUnit:{ra.get("major_unit",2)}
    numFmt: {ra.get("num_fmt","0.0")}         gridlines: None
    max={ra.get("max",20)} hard caps old P/E spikes — prevents compressing current view

  CATEGORY AXIS (bottom):
    numFmt: "{xa.get("num_fmt","MMM-YY")}"   tickLblSkip:{xa.get("tick_label_skip",26)}  tickMarkSkip:{xa.get("tick_mark_skip",26)}
    With weekly data: skip=26 ≈ every 6 months. Use skip=13 for ~3 months.
    gridlines: None

  LEGEND:
    position: "{_get(cfg,"chart","legend_position",default="t")}"   overlay: {_get(cfg,"chart","legend_overlay",default=True)}

━━━━ XML PATCH (applied after wb.save() via patch_chart_xml()) ━━━━
  File: xl/charts/chart1.xml inside the xlsx zip
  Namespaces: C_NS="http://schemas.openxmlformats.org/drawingml/2006/chart"
              A_NS="http://schemas.openxmlformats.org/drawingml/2006/main"

  PATCH 1 — Axis title boxes (both valAx elements, axId=100 and axId=200):
    <c:title>
      <c:tx> → replace with <c:rich> bold text (see _rich_bold_tx helper)
              font: Arial, sz={atb.get("font_size_pt",9)*100} (hundredths of pt), bold=1, color={atb.get("text_color","000000")}
      <c:spPr>
        <a:solidFill><a:srgbClr val="{atb.get("fill","FFFF00")}"/></a:solidFill>  ← yellow fill
        <a:ln w="{atb.get("border_width_emu",9525)}">
          <a:solidFill><a:srgbClr val="{atb.get("border_color","000000")}"/></a:solidFill>  ← black border
        </a:ln>
      </c:spPr>

  PATCH 2 — Secondary axis right-side position:
    Find <valAx> where <axId val="200">
    Set <axPos val="r"/>   (openpyxl writes "l" by default — must override)

  PATCH 3 — Legend box:
    <c:legend>
      <c:spPr>
        <a:solidFill><a:srgbClr val="{lb.get("fill","FFFFFF")}"/></a:solidFill>  ← white fill
        <a:ln w="{lb.get("border_width_emu",12700)}">
          <a:solidFill><a:srgbClr val="{lb.get("border_color","404040")}"/></a:solidFill>  ← dark grey border
        </a:ln>
      </c:spPr>

━━━━ CELL FORMATTING STANDARDS ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Font: Arial 9pt throughout
  Header cells:  fill={NAVY} (navy),  color=FFFFFF (white),  bold=True
  Input cells:   fill={YELLOW} (yellow), color={BLUE_INPUT} (blue), bold=False
  Formula cells: italic=True, color=333333, no special fill
  Row bands:     even rows fill={GREY}, odd rows fill={WHITE}
  All cells:     Border: thin, color={BORDER_COL}  |  Alignment: vertical=center

━━━━ ADDING A NEW COMPANY ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  1. Copy valuations/nhc_ax.yaml → valuations/<new_ticker_slug>.yaml
  2. Update all fields: ticker, company_name, business, stock_type
  3. Set buy_multiple, sell_multiple, rror, norm_eps (null if not cyclical)
  4. Adjust chart.left_axis.max to ~20% above expected peak price
  5. Adjust chart.right_axis.max to accommodate expected P/E range
  6. Update tick_label_skip (26 = 6 months with weekly data)
  7. Fill in earnings history (date, period, ttm_eps, ttm_div, notes)
  8. Run: python scripts/build_value_chart.py NEW.AX
     Or:  python scripts/build_value_chart.py NEW.AX --price-csv path/to/prices.csv

━━━━ PRICE DATA NOTE ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Source: {_get(cfg,"price_data","source_hint",default="marketindex.com.au")}
  Resample: df.set_index("Date").resample("{_get(cfg,"price_data","resample_day",default="W-FRI")}").last().dropna().reset_index()
  Why weekly: ~530 rows vs ~2600 daily. Thinner, cleaner price line without changing line width.
  Microsoft 365: {_get(cfg,"price_data","stockhistory_formula",default="")}

━━━━ LIVE FUNDAMENTALS (AUTO-FETCH) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  If earnings: [] is empty in the YAML config, Wally automatically fetches live
  fundamentals before building the workbook:
    - Quarterly EPS history   → Alpha Vantage EARNINGS endpoint (requires ALPHAVANTAGE_API_KEY)
    - Dividend history        → yfinance (no API key required)
    - TTM EPS                 → rolling sum of last 4 reported quarterly EPS values
    - TTM Div                 → rolling sum of dividends in the 12 months before each EPS date
    - eps_scale               → 100  (converts USD/AUD dollars to cents as stored in workbook)

  Manual YAML earnings history is NOT the only source — live data is fetched automatically
  for US stocks when ALPHAVANTAGE_API_KEY is set.

  Build status logged to stdout:
    [wally] TICKER price_source=yfinance earnings_source=alphavantage dividends_source=yfinance
    [wally] TICKER earnings_records=N dividend_records=M
    [wally] TICKER build_status=FULL_BUILD | PRICE_ONLY_BUILD

  If live fundamentals cannot be fetched, the workbook is clearly labelled PRICE_ONLY with
  a visible warning in the EarningsData sheet — blank charts are never silently produced.
"""

    for i, line in enumerate(prompt.splitlines(), start=2):
        safe = ("'" + line) if line.startswith("=") else line
        c = ws5.cell(i, 1, safe)
        c.font = Font(name="Courier New", size=9)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PUBLIC FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
#  MATPLOTLIB PNG CHART
# ─────────────────────────────────────────────────────────────────────────────

def build_chart_png(
    cfg: dict,
    price_df: pd.DataFrame,
    output_path: str,
) -> str | None:
    """
    Render a dual-axis matplotlib PNG chart equivalent to the Excel ValueChart sheet.
    Returns the saved path, or None if matplotlib is unavailable.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")          # non-interactive backend — safe in CI / GitHub Actions
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        print("[vcb] matplotlib not available — skipping PNG generation")
        return None

    ticker = cfg["ticker"].split(".")[0]
    name   = cfg["company_name"]
    bmult  = cfg["buy_multiple"]
    smult  = cfg.get("sell_multiple")
    rror   = float(cfg["rror"])
    norm   = cfg.get("norm_eps")

    la  = _get(cfg, "chart", "left_axis")  or {}
    ra  = _get(cfg, "chart", "right_axis") or {}
    lns = _get(cfg, "chart", "lines")      or {}

    def _hex(key: str, fallback: str) -> str:
        return "#" + lns.get(key, {}).get("color", fallback)

    # ── Compute value series ──────────────────────────────────────────────────
    earn_dates = [
        pd.Timestamp(e["date"] if isinstance(e["date"], date)
                     else datetime.strptime(str(e["date"]), "%Y-%m-%d").date())
        for e in cfg["earnings"]
    ]
    earn_eps = [float(e["ttm_eps"]) for e in cfg["earnings"]]
    earn_div = [float(e["ttm_div"]) for e in cfg["earnings"]]

    dates, prices, v_buys, v_sells, drrors, raw_pe = [], [], [], [], [], []

    for _, row in price_df.iterrows():
        d     = pd.Timestamp(row["Date"])
        close = float(row["Close"])
        eps   = _lookup_last(d, earn_dates, earn_eps)
        div   = _lookup_last(d, earn_dates, earn_div)

        dates.append(d.to_pydatetime())
        prices.append(close)
        v_buys.append((eps * bmult / 100)   if eps is not None else None)
        v_sells.append((eps * smult / 100)  if (smult and eps is not None) else None)
        drrors.append((div / rror / 100)    if (div is not None and rror > 0) else None)
        raw_pe.append((close / (eps / 100)) if (eps is not None and eps > 0) else np.nan)

    pe_smooth = (pd.Series(raw_pe, dtype=float)
                   .rolling(window=12, min_periods=4).mean().tolist())

    # ── Plot ──────────────────────────────────────────────────────────────────
    fig, ax1 = plt.subplots(figsize=(14, 7))
    fig.patch.set_facecolor("#FFFFFF")
    ax1.set_facecolor("#FFFFFF")

    ax1.plot(dates, prices, color=_hex("price", "1565C0"),
             linewidth=1.0, label="Price (AUD $)", zorder=3)

    ax2 = ax1.twinx()

    ax2.plot(dates, v_buys, color=_hex("value_buy", "1A5E20"),
             linewidth=2.0, linestyle="--", label=f"Value Buy ({bmult}× EPS)", zorder=2)
    if smult:
        ax2.plot(dates, v_sells, color=_hex("value_sell", "C0392B"),
                 linewidth=2.0, linestyle="--", label=f"Value Sell ({smult}× EPS)", zorder=2)
    ax2.plot(dates, drrors, color=_hex("div_rror", "2E7D32"),
             linewidth=2.0, linestyle="--", label=f"Div/RRoR ({int(rror*100)}%)", zorder=2)
    ax2.plot(dates, pe_smooth, color=_hex("pe_smooth", "B8860B"),
             linewidth=1.5, linestyle="-", label="P/E (12wk avg)", zorder=2)

    # Axis limits
    ax1.set_ylim(float(la.get("min", 0.0)), float(la.get("max", 10.0)))
    ax2.set_ylim(float(ra.get("min", 0.0)), float(ra.get("max", 20.0)))

    # Axis labels — yellow box style matching the Excel chart
    _label_kw = dict(fontsize=10, fontweight="bold",
                     bbox=dict(boxstyle="round,pad=0.3",
                                facecolor="#FFFF00", edgecolor="#000000", linewidth=1))
    ax1.set_ylabel(la.get("label", "Price (AUD $)"), **_label_kw)
    ax2.set_ylabel(ra.get("label", "P/E Ratio"),     **_label_kw)

    # Title — navy banner
    _exchange_png = cfg.get("exchange", "")
    _exl_png = f"{_exchange_png}: " if _exchange_png else ""
    title = (f"{name} ({_exl_png}{ticker})  —  Buy / Sell Zone Chart  |  Buy {bmult}×"
             + (f"  |  Sell {smult}×" if smult else ""))
    if norm:
        title += (f"\nNorm EPS: {norm}¢  →  Buy ${norm*bmult/100:.2f}"
                  + (f"  /  Sell ${norm*smult/100:.2f}" if smult else ""))
    ax1.set_title(title, fontsize=11, fontweight="bold",
                  color="white", backgroundcolor="#1F2D4E",
                  pad=8, loc="left")

    # X axis
    ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%b-%y"))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right", fontsize=8)

    # Gridlines off
    ax1.grid(False)
    ax2.grid(False)

    # Legend — combined, white box with border
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2,
               loc="upper left", fontsize=8, frameon=True,
               framealpha=1.0, edgecolor="#404040")

    plt.tight_layout()
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out), dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"[vcb] PNG saved: {out}")
    return str(out)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN PUBLIC FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def build_value_chart(
    ticker_or_config_path: str,
    output_path: str | None = None,
    price_csv_path: str | None = None,
    drive_folder_id: str | None = None,
    save_png: bool = False,
) -> str:
    """
    Build the canonical 5-sheet value chart workbook for any company on any market.

    Produces workbook sheets (in order): Settings, EarningsData, PriceData,
    ValueChart, FuturePrompt.  If no config file exists for *ticker_or_config_path*,
    a starter config is auto-created and the workbook is built with DATA REQUIRED
    placeholders — no FileNotFoundError is raised.

    Args:
        ticker_or_config_path: Ticker (e.g. "NHC.AX", "POOL", "2914.T") or
                               full path to a YAML config file.
        output_path:           Where to save the xlsx.  Defaults to
                               ``outputs/<TICKER>_value_chart.xlsx``.
        price_csv_path:        Optional path to a price CSV (Date as YYYYMMDD,
                               Close column).  Falls back to yfinance then
                               synthetic placeholder.
        drive_folder_id:       If set, uploads/replaces file in Google Drive.
        save_png:              If True, also saves a PNG chart alongside the xlsx.

    Returns:
        str: Local path to saved xlsx, or Drive URL when *drive_folder_id* is set.
    """
    cfg = load_config(ticker_or_config_path)

    # Auto-fetch live fundamentals when earnings history is absent from config.
    # This populates cfg["earnings"] from Alpha Vantage (EPS) + yfinance (dividends)
    # so that TII75 US stocks build with real value lines instead of blank DATA REQUIRED.
    if not cfg.get("earnings"):
        live_earnings, _, _, price_only_warning = _fetch_live_fundamentals(
            cfg["ticker"], cfg
        )
        if live_earnings:
            cfg["earnings"] = live_earnings
        elif price_only_warning:
            cfg["_price_only_warning"] = price_only_warning

    if output_path is None:
        out_dir = Path("outputs")
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(out_dir / f"{_ticker_slug(cfg['ticker'])}_value_chart.xlsx")

    price_df = get_price_data(cfg, price_csv_path)

    wb = Workbook()
    wb.remove(wb.active)

    _build_settings(wb, cfg)
    _build_earnings(wb, cfg)
    _build_price_data(wb, cfg, price_df)
    _build_chart(wb, cfg, wb["PriceData"])
    _build_future_prompt(wb, cfg)

    wb.active = wb["ValueChart"]
    wb["EarningsData"].freeze_panes = "A3"
    wb["PriceData"].freeze_panes    = "A3"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    patch_chart_xml(out, cfg)

    kb = round(out.stat().st_size / 1024, 1)
    print(f"[vcb] Saved & patched: {out}  ({kb} KB)")

    if save_png:
        png_path = str(out.with_suffix(".png"))
        build_chart_png(cfg, price_df, png_path)

    if drive_folder_id:
        try:
            from wally.drive_upload import upload_or_replace_xlsx
            url = upload_or_replace_xlsx(
                out,
                drive_name=cfg["ticker"].split(".")[0],
                folder_id=drive_folder_id,
            )
            print(f"[vcb] Uploaded to Drive: {url}")
            return url
        except Exception as e:
            print(f"[vcb] Drive upload failed: {e}")

    return str(out)
